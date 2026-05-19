"""
validadores.py — Validações do pipeline de Antilhas.

Responsabilidades:
  - Validar nome do arquivo enviado pela loja
  - Validar estrutura interna da planilha
  - Validar consistência da data interna vs nome do arquivo
"""

import re
import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# Regex esperada: LOJA_<codigo>_<AAAA-MM-DD>.xlsx ou .xlsm
_PADRAO_NOME = re.compile(
    r"^LOJA_(?P<codigo>\d+)_(?P<ano>\d{4})-(?P<mes>\d{2})-(?P<dia>\d{2})\.(xlsx|xlsm)$",
    re.IGNORECASE,
)

ABA_CONTAGEM = "CONTAGEM Semanal"


class ErroValidacao(Exception):
    """Erro esperado de validação — arquivo vai para quarentena."""


def validar_nome_arquivo(nome_arquivo: str, codigos_validos: set[str]) -> tuple[str, date]:
    """
    Valida o nome do arquivo e retorna (codigo_loja, data_contagem).

    Raises:
        ErroValidacao: se o nome não bate com o padrão, código não existe ou data é inválida.
    """
    match = _PADRAO_NOME.match(nome_arquivo)
    if not match:
        raise ErroValidacao(
            f"Nome fora do padrão esperado 'LOJA_<codigo>_<AAAA-MM-DD>.xlsx': '{nome_arquivo}'"
        )

    codigo = match.group("codigo")
    if codigo not in codigos_validos:
        raise ErroValidacao(
            f"Código de loja '{codigo}' não encontrado no config.yaml."
        )

    try:
        data_nome = date(
            int(match.group("ano")),
            int(match.group("mes")),
            int(match.group("dia")),
        )
    except ValueError as exc:
        raise ErroValidacao(f"Data inválida no nome do arquivo: {exc}") from exc

    logger.debug("Nome válido: código=%s data=%s", codigo, data_nome)
    return codigo, data_nome


def validar_estrutura_planilha(caminho: Path) -> openpyxl.Workbook:
    """
    Abre a planilha com data_only=True e valida:
      - Aba 'CONTAGEM Semanal' existe
      - A1 contém 'RESPONSÁVEL:'
      - D3 contém 'DATA CONTAGEM'
      - E3 é um datetime

    Retorna o Workbook aberto para reuso.

    Raises:
        ErroValidacao: qualquer divergência estrutural.
    """
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
    except Exception as exc:
        raise ErroValidacao(f"Não foi possível abrir o arquivo: {exc}") from exc

    if ABA_CONTAGEM not in wb.sheetnames:
        raise ErroValidacao(
            f"Aba '{ABA_CONTAGEM}' não encontrada. Abas presentes: {wb.sheetnames}"
        )

    ws = wb[ABA_CONTAGEM]

    val_a1 = _texto(ws["A1"].value)
    if "RESPONSÁVEL:" not in val_a1:
        raise ErroValidacao(
            f"A1 deveria conter 'RESPONSÁVEL:', encontrado: '{ws['A1'].value}'"
        )

    val_d3 = _texto(ws["D3"].value)
    if "DATA CONTAGEM" not in val_d3:
        raise ErroValidacao(
            f"D3 deveria conter 'DATA CONTAGEM', encontrado: '{ws['D3'].value}'"
        )

    val_e3 = ws["E3"].value
    if not isinstance(val_e3, (datetime, date)):
        raise ErroValidacao(
            f"E3 deveria ser uma data/datetime, encontrado: '{val_e3}' ({type(val_e3).__name__})"
        )

    logger.debug("Estrutura da planilha válida: %s", caminho.name)
    return wb


def validar_data_interna(wb: openpyxl.Workbook, data_nome: date) -> date:
    """
    Compara a data em E3 da planilha com a data extraída do nome do arquivo.

    Retorna a data confirmada.

    Raises:
        ErroValidacao: se as datas divergirem.
    """
    ws = wb[ABA_CONTAGEM]
    val_e3 = ws["E3"].value

    # Normalizar para date independente de ser datetime ou date
    if isinstance(val_e3, datetime):
        data_interna = val_e3.date()
    else:
        data_interna = val_e3

    if data_interna != data_nome:
        raise ErroValidacao(
            f"Data interna ({data_interna}) difere da data no nome do arquivo ({data_nome})."
        )

    logger.debug("Data interna confirmada: %s", data_interna)
    return data_interna


def extrair_responsavel(wb: openpyxl.Workbook) -> str:
    """Retorna o nome do responsável da célula B1."""
    ws = wb[ABA_CONTAGEM]
    return _texto(ws["B1"].value) or "Não informado"


# ── Auxiliares ──────────────────────────────────────────────────────────────

def _texto(valor) -> str:
    """Converte qualquer valor para string maiúscula sem espaços extras."""
    if valor is None:
        return ""
    return str(valor).strip().upper()
