"""
processadores/lojas.py — Extração de dados do arquivo enviado pela loja.

Layout da aba 'CONTAGEM Semanal':
  Bloco esquerdo  → colunas A(produto) B(pct) D(fechadas) E(abertas) F(total)
  Bloco direito   → colunas H(produto) I(pct) K(fechadas) L(abertas) M(total)

Retorna um dicionário:
  {
    "SACOLA PP": 42,
    "PAPEL DE SEDA": 10,
    ...
  }
onde as chaves são os nomes uppercase como estão no config (nome_loja).
"""

import logging
from datetime import date
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

ABA_CONTAGEM = "CONTAGEM Semanal"

# Índices de coluna (base 1, como openpyxl usa com ws.cell(row, col))
_COL = {
    # Bloco esquerdo
    "esq_produto":   1,   # A
    "esq_fechadas":  4,   # D
    "esq_abertas":   5,   # E
    "esq_total":     6,   # F
    # Bloco direito
    "dir_produto":   8,   # H
    "dir_fechadas":  11,  # K
    "dir_abertas":   12,  # L
    "dir_total":     13,  # M
}


def extrair_dados_loja(
    wb: openpyxl.Workbook,
    mapa_linhas: dict,
    mapa_produtos: list[dict],
) -> dict[str, Any]:
    """
    Extrai os valores de todos os produtos configurados no config.yaml.

    Args:
        wb:            Workbook já aberto com data_only=True.
        mapa_linhas:   Seção 'linhas_arquivo' do config (bloco_esquerdo + bloco_direito).
        mapa_produtos: Lista de produtos do config (seção 'produtos').

    Returns:
        Dicionário {nome_loja_upper: valor_numerico}.
        Produtos sem valor encontrado ficam com None.
    """
    ws = wb[ABA_CONTAGEM]
    resultado: dict[str, Any] = {}

    linhas_esq = mapa_linhas.get("bloco_esquerdo", {})
    linhas_dir = mapa_linhas.get("bloco_direito", {})

    for produto in mapa_produtos:
        nome_loja: str = produto["nome_loja"].strip().upper()
        bloco: str = produto["bloco"]
        tipo_valor: str = produto["valor"]

        try:
            if bloco == "esquerdo":
                valor = _ler_bloco_esquerdo(ws, nome_loja, tipo_valor, linhas_esq)
            elif bloco == "direito":
                valor = _ler_bloco_direito(ws, nome_loja, tipo_valor, linhas_dir)
            else:
                logger.warning("Bloco desconhecido '%s' para produto '%s'.", bloco, nome_loja)
                valor = None

            resultado[nome_loja] = valor
            logger.debug("Produto '%s' → %s", nome_loja, valor)

        except Exception as exc:
            logger.warning("Erro ao ler produto '%s': %s", nome_loja, exc)
            resultado[nome_loja] = None

    return resultado


# ── Leitores de bloco ───────────────────────────────────────────────────────

def _ler_bloco_esquerdo(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    nome_produto: str,
    tipo_valor: str,
    linhas: dict,
) -> Any:
    """Lê o valor de um produto no bloco esquerdo (cols A-F)."""
    linha = _linha_do_produto(nome_produto, linhas)
    return _resolver_valor(
        ws=ws,
        linha=linha,
        col_fechadas=_COL["esq_fechadas"],
        col_abertas=_COL["esq_abertas"],
        col_total=_COL["esq_total"],
        tipo_valor=tipo_valor,
    )


def _ler_bloco_direito(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    nome_produto: str,
    tipo_valor: str,
    linhas: dict,
) -> Any:
    """Lê o valor de um produto no bloco direito (cols H-M)."""
    linha = _linha_do_produto(nome_produto, linhas)
    return _resolver_valor(
        ws=ws,
        linha=linha,
        col_fechadas=_COL["dir_fechadas"],
        col_abertas=_COL["dir_abertas"],
        col_total=_COL["dir_total"],
        tipo_valor=tipo_valor,
    )


def _linha_do_produto(nome_produto: str, linhas: dict) -> int:
    """
    Retorna o número de linha (base 1) para o produto.
    Busca com .strip().upper() em ambos os lados para tolerar espaços.
    """
    nome_upper = nome_produto.strip().upper()
    for chave, linha in linhas.items():
        if chave.strip().upper() == nome_upper:
            return int(linha)
    raise ValueError(f"Linha não configurada para o produto '{nome_produto}'.")


def _resolver_valor(
    ws,
    linha: int,
    col_fechadas: int,
    col_abertas: int,
    col_total: int,
    tipo_valor: str,
) -> Any:
    """
    Retorna o valor numérico conforme tipo_valor:
      "total"    → lê col_total; se None, soma fechadas+abertas
      "fechadas" → lê col_fechadas
      "abertas"  → lê col_abertas
    """
    if tipo_valor == "total":
        total = _numero(ws.cell(row=linha, column=col_total).value)
        if total is None:
            fechadas = _numero(ws.cell(row=linha, column=col_fechadas).value) or 0
            abertas  = _numero(ws.cell(row=linha, column=col_abertas).value) or 0
            total = fechadas + abertas
            logger.debug(
                "Total calculado manualmente (linha %d): %d + %d = %d",
                linha, fechadas, abertas, total,
            )
        return total

    if tipo_valor == "fechadas":
        return _numero(ws.cell(row=linha, column=col_fechadas).value)

    if tipo_valor == "abertas":
        return _numero(ws.cell(row=linha, column=col_abertas).value)

    raise ValueError(f"tipo_valor desconhecido: '{tipo_valor}'")


# ── Auxiliares ──────────────────────────────────────────────────────────────

def _numero(valor: Any) -> Any:
    """
    Converte o valor da célula para int ou float.
    Retorna None se o valor for None, string vazia ou não numérico.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return valor
    texto = str(valor).strip()
    if not texto:
        return None
    try:
        # Preferir int quando possível (contagens são inteiras)
        float_val = float(texto.replace(",", "."))
        return int(float_val) if float_val == int(float_val) else float_val
    except ValueError:
        return None
