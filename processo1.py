"""
Processo 1 — Automação de inserção de coluna e preenchimento das Planilhas Pai.
Replicação em Python do macro VBA de controle de estoque Antilhas.

Uso:
    python processo1.py                    # execução normal (todas as regiões)
    python processo1.py --dry-run          # simula sem escrever nem fazer backup
    python processo1.py --regiao praia     # processa só a Praia
    python processo1.py --regiao jau       # processa só Jaú
    python processo1.py --regiao bauru     # processa só Bauru
"""
from __future__ import annotations

import argparse
import copy
import datetime as dt
import logging
import shutil
import sys
from pathlib import Path
from typing import Callable, Dict, Optional

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
import yaml


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def vz(valor) -> float:
    """Retorna 0 se None, não-numérico ou negativo. Caso contrário retorna float."""
    try:
        v = float(valor)
        return v if v >= 0 else 0
    except (TypeError, ValueError):
        return 0


def copiar_formato(ws, col_origem: int, col_destino: int) -> None:
    """Copia estilos célula a célula de col_origem para col_destino (sem mexer em larguras)."""
    for row_idx in range(1, ws.max_row + 1):
        src = ws.cell(row=row_idx, column=col_origem)
        dst = ws.cell(row=row_idx, column=col_destino)
        if src.has_style:
            try:
                dst.font          = copy.copy(src.font)
                dst.fill          = copy.copy(src.fill)
                dst.border        = copy.copy(src.border)
                dst.alignment     = copy.copy(src.alignment)
                dst.number_format = src.number_format
            except Exception:
                pass  # pula células mescladas ou com proteção


# ─────────────────────────────────────────────────────────────────────────────
# BACKUP
# ─────────────────────────────────────────────────────────────────────────────

def backup(cfg: dict, dry_run: bool, logger: logging.Logger) -> Path:
    """Copia os 3 .xlsm para 01_BACKUP/<AAAA-MM-DD_HHhMM>/ antes de qualquer modificação."""
    agora = dt.datetime.today()
    nome_pasta = agora.strftime("%Y-%m-%d_%Hh%M")
    pasta_backup = Path(cfg["dirs"]["backup"]) / nome_pasta

    if dry_run:
        logger.info("[DRY-RUN] Backup seria criado em: %s", pasta_backup)
        print(f"  [DRY-RUN] Backup seria criado em: {pasta_backup}")
        return pasta_backup

    try:
        pasta_backup.mkdir(parents=True, exist_ok=True)
        for regiao, caminho in cfg["planilhas_pai"].items():
            src = Path(caminho)
            if not src.exists():
                raise FileNotFoundError(f"Planilha não encontrada: {src}")
            dst = pasta_backup / src.name
            shutil.copy2(src, dst)
            logger.debug("  Backup: %s → %s", src.name, dst)
        logger.info("Backup criado em: %s", pasta_backup)
        print(f"  Backup criado em: {pasta_backup}")
        return pasta_backup
    except Exception as exc:
        logger.critical("Falha no backup: %s", exc)
        sys.exit(f"ERRO CRÍTICO: Backup falhou — {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# LEITOR DAS PLANILHAS AUXILIARES (busca por código de produto + código de loja)
# ─────────────────────────────────────────────────────────────────────────────

class LeitorAuxiliar:
    """
    Indexa AUXILIAR.xlsx / AUXILIAR_VDS.xlsx por:
      - Código do produto (coluna A: '00000000089402 - CAIXA...' → '00000000089402')
      - Código da loja    (linha 1:  '12973', '24790', etc.)

    Uso:
        leitor = LeitorAuxiliar(wsAux, "AUXILIAR")
        valor = leitor.get("00000000089402", " PL 12973")
                       ↑ código do produto    ↑ nome da aba destino (extrai '12973')

    Se o produto ou a loja não forem encontrados, retorna 0 e registra um aviso
    em self.faltantes (impresso no final pelo main).
    """

    def __init__(self, ws, nome: str = "AUXILIAR"):
        self.ws = ws
        self.nome = nome
        self.produtos: Dict[str, int] = {}   # codigo_produto → linha
        self.lojas: Dict[str, int] = {}      # codigo_loja    → coluna
        self.faltantes: set = set()
        self._construir_indice()

    def _construir_indice(self) -> None:
        # Produtos (coluna A, linhas 2+)
        for row in range(2, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=1).value
            if val is None:
                continue
            s = str(val).strip()
            if not s or s.lower() in ("total", "produtos"):
                continue
            # Código fica antes de " - "
            cod = s.split(" - ")[0].strip()
            if cod and cod[0].isdigit():
                # Se duplicar, mantém a primeira ocorrência
                if cod not in self.produtos:
                    self.produtos[cod] = row

        # Lojas (linha 1, colunas 2+)
        for col in range(2, self.ws.max_column + 1):
            val = self.ws.cell(row=1, column=col).value
            if val is None:
                continue
            cod = str(val).strip()
            if cod and cod.lower() != "produtos":
                self.lojas[cod] = col

    @staticmethod
    def _extrair_codigo_loja(nome_aba: str) -> str:
        """' PL 12973' → '12973'  |  'ATC - 23012' → '23012'  |  'ER BOQ - 23614' → '23614'"""
        partes = nome_aba.strip().split()
        return partes[-1] if partes else ""

    def get(self, produto_codigo: str, nome_aba_destino: str):
        """Retorna o valor da célula (produto × loja). Retorna 0 se não encontrar."""
        loja_codigo = self._extrair_codigo_loja(nome_aba_destino)
        row = self.produtos.get(produto_codigo)
        col = self.lojas.get(loja_codigo)
        if row is None:
            self.faltantes.add(
                f"{self.nome}: produto '{produto_codigo}' não encontrado "
                f"(usado pela aba '{nome_aba_destino}')"
            )
            return 0
        if col is None:
            self.faltantes.add(
                f"{self.nome}: loja '{loja_codigo}' não encontrada na linha 1 "
                f"(aba destino: '{nome_aba_destino}')"
            )
            return 0
        return self.ws.cell(row=row, column=col).value


# ─────────────────────────────────────────────────────────────────────────────
# ELEGIBILIDADE DE ABAS
# ─────────────────────────────────────────────────────────────────────────────

_PREFIXOS_IGNORADOS = ("__", "TOTAL")
_NOMES_IGNORADOS = frozenset({
    "Planilha1", "AUXILIAR", "AUXILIAR VD'S",
    "LOJAS JAÚ", "LOJAS PRAIA", "LOJAS BAURU",   # tratadas por avancar_referencias_lojas
})
_ABAS_LOJAS = frozenset({"LOJAS JAÚ", "LOJAS PRAIA", "LOJAS BAURU"})


def eh_elegivel(nome_aba: str) -> bool:
    if nome_aba in _NOMES_IGNORADOS:
        return False
    return not any(nome_aba.startswith(p) for p in _PREFIXOS_IGNORADOS)


# ─────────────────────────────────────────────────────────────────────────────
# INSERÇÃO DE COLUNA (replicar ProcessarAba do VBA)
# ─────────────────────────────────────────────────────────────────────────────

def _cel_tem_data(cell) -> bool:
    """Retorna True se a célula contém uma data — via tipo Python, número, fórmula ou texto DD/MM."""
    val = cell.value
    if val is None:
        return False
    if isinstance(val, (dt.date, dt.datetime)):
        return True
    fmt = (cell.number_format or "").lower()
    _DATE_PATTERNS = ("d/m", "dd/mm", "d-m", "m/d", "mm/dd", "dd-mm")
    if isinstance(val, (int, float)) and val > 0:
        if any(p in fmt for p in _DATE_PATTERNS):
            return True
    if isinstance(val, str) and val.startswith("="):
        if any(p in fmt for p in _DATE_PATTERNS):
            return True
    # Texto literal no formato DD/MM (ex: "14/05", "07/05") — comum em cabeçalhos
    if isinstance(val, str) and not val.startswith("="):
        partes = val.strip().split("/")
        if len(partes) == 2 and partes[0].strip().isdigit() and partes[1].strip().isdigit():
            try:
                if 1 <= int(partes[0].strip()) <= 31 and 1 <= int(partes[1].strip()) <= 12:
                    return True
            except ValueError:
                pass
    return False


def _encontrar_todas_col_usou(ws) -> list:
    """Retorna lista com índices 1-based de TODAS as colunas que contêm 'USOU' (linhas 1-30)."""
    resultado = []
    for col in range(1, ws.max_column + 1):
        for row in range(1, 31):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip().upper() == "USOU":
                resultado.append(col)
                break  # encontrou USOU nessa coluna — passa para a próxima
    return resultado


def inserir_coluna(ws, dry_run: bool, logger: logging.Logger, nome_aba: str) -> Optional[Dict[int, int]]:
    """
    Localiza TODAS as colunas USOU e insere 1 coluna à esquerda de cada uma,
    processando da direita para a esquerda (igual ao VBA) para evitar deslocamentos.
    Após as inserções:
      - Escreve data de hoje em TODAS as linhas com data na coluna anterior (Fix 2)
      - Reescreve a fórmula da coluna USOU com referências corretas (Fix 1)
      - Restaura todas as larguras de coluna a partir de snapshot tirado antes (Fix 3)

    Retorna um dicionário {índice_usou: posição_da_nova_coluna} (1-based, em ordem
    da esquerda para a direita) ou None se nenhum USOU for encontrado.
        Exemplo: {1: 28, 2: 60}  → USOU 1 nova col em 28, USOU 2 nova col em 60.
    """
    cols_usou = _encontrar_todas_col_usou(ws)
    if not cols_usou:
        logger.warning("     Aba '%s': coluna USOU não encontrada — pulada", nome_aba)
        return None

    sorted_usou_orig = sorted(cols_usou)
    # Após inserir colunas da direita para a esquerda, a nova coluna do i-ésimo
    # USOU (1-based, esquerda→direita) fica em (orig + (i - 1)).
    cols_novas_por_usou: Dict[int, int] = {
        i: orig + (i - 1) for i, orig in enumerate(sorted_usou_orig, start=1)
    }

    if dry_run:
        logger.debug("    [DRY-RUN] Aba '%s': %d USOU(s) em %s → novas cols %s",
                     nome_aba, len(cols_usou), cols_usou, cols_novas_por_usou)
        return cols_novas_por_usou

    ultima_linha = ws.max_row
    sorted_usou = sorted(cols_usou)  # ordem crescente (para cálculos de deslocamento)

    # ── Fix 3 (parte 1): salvar larguras E estado oculto antes de qualquer inserção
    # Itera sobre column_dimensions diretamente para capturar TODAS as colunas
    # com configuração explícita — inclusive colunas ocultas sem dados.
    all_widths: dict = {}
    all_hidden: dict = {}
    for letra, dim in ws.column_dimensions.items():
        col_num = column_index_from_string(letra)
        if dim.width and dim.width > 0:
            all_widths[col_num] = dim.width
        if dim.hidden:
            all_hidden[col_num] = True

    # ── Inserir colunas da direita para a esquerda (igual ao VBA) ───────────────
    for col_usou in sorted(cols_usou, reverse=True):
        col_nova = col_usou

        ws.insert_cols(col_nova)

        # Copiar estilos de célula da coluna anterior (sem largura — gerenciada abaixo)
        if col_nova > 1:
            copiar_formato(ws, col_nova - 1, col_nova)

        # Fix 2: escrever data em TODAS as linhas que têm data na coluna anterior
        if col_nova > 1:
            for row in range(1, ultima_linha + 1):
                if _cel_tem_data(ws.cell(row=row, column=col_nova - 1)):
                    cell = ws.cell(row=row, column=col_nova)
                    cell.value = dt.datetime.today()
                    cell.number_format = "DD/MM"

    # ── Fix 1: reescrever fórmulas USOU com referências ajustadas ───────────────
    for orig_usou in sorted_usou:
        # Cada USOU original em 'orig_usou' deslocou para frente 1 vez por cada
        # nova coluna inserida à esquerda ou na mesma posição
        n_shifts = sum(1 for u in sorted_usou if u <= orig_usou)
        final_col = orig_usou + n_shifts  # posição final da coluna USOU

        if final_col < 3:
            continue  # precisaria de pelo menos 2 colunas à esquerda para a fórmula

        letra_prev2 = get_column_letter(final_col - 2)
        letra_prev1 = get_column_letter(final_col - 1)
        letra_self  = get_column_letter(final_col)

        for row in range(1, ultima_linha + 1):
            cell = ws.cell(row=row, column=final_col)
            if (cell.value is not None
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")):
                cell.value = (
                    f"={letra_prev2}{row}"
                    f"+IF(NOT(ISBLANK(OFFSET({letra_self}{row},0,6))),"
                    f"N(OFFSET({letra_self}{row},0,6)),"
                    f"N(OFFSET({letra_self}{row},0,7)))"
                    f"-{letra_prev1}{row}"
                )

    # ── Fix 3 (parte 2): limpar e reconstruir larguras + visibilidade ───────────
    ws.column_dimensions.clear()

    # Colunas originais: restaurar largura e oculto na nova posição
    # Usa o conjunto exato de colunas que tinham configuração explícita
    all_orig_cols = set(all_widths.keys()) | set(all_hidden.keys())
    for orig_col in all_orig_cols:
        n_shifts = sum(1 for u in sorted_usou if u <= orig_col)
        nova_letra = get_column_letter(orig_col + n_shifts)
        if orig_col in all_widths:
            ws.column_dimensions[nova_letra].width = all_widths[orig_col]
        if orig_col in all_hidden:
            ws.column_dimensions[nova_letra].hidden = True

    # Colunas novas inseridas: herdam largura da coluna à esquerda; nunca ocultas
    for u in sorted_usou:
        n_shifts_antes = sum(1 for other_u in sorted_usou if other_u < u)
        new_pos = u + n_shifts_antes          # posição da nova coluna
        vizinha_orig = u - 1                  # coluna original à esquerda
        if vizinha_orig in all_widths:
            ws.column_dimensions[get_column_letter(new_pos)].width = all_widths[vizinha_orig]

    # ── Ocultar semanas antigas: manter só as 2 mais recentes por USOU ──────────
    # Para cada USOU, escaneia as colunas de data contíguas à sua esquerda
    # (mais recente → mais antiga) e oculta tudo além das 2 primeiras.
    for orig_usou in sorted_usou:
        n_shifts = sum(1 for u in sorted_usou if u <= orig_usou)
        final_usou_col = orig_usou + n_shifts

        # Coletar bloco contíguo de colunas de data imediatamente à esquerda do USOU.
        # Se encontrar a coluna âncora "Novo Pedido" (não é data), pula e continua
        # escaneando para a esquerda — isso permite alcançar o histórico de datas que
        # fica à esquerda do Novo Pedido em planilhas com esse layout.
        date_cols_region: list = []
        for col in range(final_usou_col - 1, 0, -1):
            tem_data = any(
                _cel_tem_data(ws.cell(row=row, column=col))
                for row in range(1, 16)
            )
            if tem_data:
                date_cols_region.append(col)
            else:
                # Verifica se é a coluna âncora "Novo Pedido" — se for, pula e continua
                is_novo_pedido = any(
                    ws.cell(row=row, column=col).value is not None
                    and str(ws.cell(row=row, column=col).value).strip().upper() == "NOVO PEDIDO"
                    for row in range(1, 16)
                )
                if not is_novo_pedido:
                    break  # coluna sem data e sem âncora Novo Pedido — fim do bloco

        # date_cols_region[0] = semana atual, [1] = semana anterior, [2+] = antigas
        for i, col in enumerate(date_cols_region):
            letra = get_column_letter(col)
            if i < 2:
                ws.column_dimensions[letra].hidden = False  # semana atual e anterior visíveis
            else:
                ws.column_dimensions[letra].hidden = True   # semanas antigas ocultas

    logger.debug("    Aba '%s': %d col(s) inserida(s), posições por USOU: %s",
                 nome_aba, len(cols_usou), cols_novas_por_usou)
    return cols_novas_por_usou


# ─────────────────────────────────────────────────────────────────────────────
# PREENCHIMENTO — PRAIA
# ─────────────────────────────────────────────────────────────────────────────

def preencher_praia(
    wb,
    leitor_aux: "LeitorAuxiliar",
    leitor_vds: "LeitorAuxiliar",
    col_novas: Dict[str, Dict[int, int]],
    dry_run: bool,
    logger: logging.Logger,
) -> Dict[str, int]:
    """
    Escreve valores nas abas da Planilha Praia lendo de AUXILIAR e AUXILIAR VD'S.
    Acesso via código de produto + código de loja (imune a inserção/reordenação
    de linhas/colunas nas auxiliares).
    Retorna {nome_aba: quantidade_de_valores_escritos}.
    """
    contagem: Dict[str, int] = {}
    _warned: set = set()

    # Acessores curtos para legibilidade
    def aux(cod_prod: str, nome_aba: str):
        return leitor_aux.get(cod_prod, nome_aba)

    def vds(cod_prod: str, nome_aba: str):
        return leitor_vds.get(cod_prod, nome_aba)

    def esc(nome_aba: str, linha: int, valor: float, usou: int = 1) -> None:
        """Escreve `valor` em (linha, coluna USOU[usou]) da aba.
        usou=1 (default) → primeiro USOU (esquerda); usou=2 → segundo USOU; etc.
        """
        if nome_aba not in col_novas:
            if nome_aba not in _warned:
                if nome_aba in wb.sheetnames:
                    logger.warning("     '%s' sem coluna USOU — não preenchida", nome_aba)
                else:
                    logger.warning("     '%s' não existe na planilha — ignorada", nome_aba)
                _warned.add(nome_aba)
            return
        cols_usou = col_novas[nome_aba]
        if usou not in cols_usou:
            # Loja não tem esse USOU — silenciosamente ignora (não é erro)
            return
        if not dry_run:
            wb[nome_aba].cell(row=linha, column=cols_usou[usou]).value = valor
        contagem[nome_aba] = contagem.get(nome_aba, 0) + 1

    # ── PL 12973 ──────────────────────────────────────────────────────────────
    n = " PL 12973"  # aba tem espaço na frente no arquivo
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 25, vz(aux("00000000086027", n)))
    esc(n, 26, vz(aux("00000000087691", n)))
    esc(n, 27, vz(aux("00000000087694", n)))
    esc(n, 30, vz(aux("00000000094778", n)))
    esc(n, 31, vz(aux("00000000094778", n)))  # mesmo valor da linha 30
    esc(n, 32, vz(aux("00000000094779", n)))
    esc(n, 33, vz(aux("00000000095151", n)))

    # ── BOQ 11734 ─────────────────────────────────────────────────────────────
    n = "BOQ 11734"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 32, vz(aux("00000000087691", n)))
    esc(n, 33, vz(aux("00000000087694", n)))
    esc(n, 45, vz(aux("00000000094775", n)))
    esc(n, 46, vz(aux("00000000094776", n)))
    esc(n, 47, vz(aux("00000000094777", n)))

    # ── TP 14462 ──────────────────────────────────────────────────────────────
    n = "TP 14462"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 25, vz(aux("00000000086027", n)))
    esc(n, 31, vz(aux("00000000087691", n)))
    esc(n, 32, vz(aux("00000000087694", n)))
    esc(n, 45, vz(aux("00000000094778", n)))

    # ── PB 5418 ───────────────────────────────────────────────────────────────
    n = "PB 5418"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 43, vz(aux("00000000094776", n)))
    esc(n, 44, vz(aux("00000000094777", n)))
    esc(n, 46, vz(aux("00000000094778", n)))

    # ── MG 11733 ──────────────────────────────────────────────────────────────
    n = "MG 11733"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 25, vz(aux("00000000086027", n)))
    esc(n, 30, vz(aux("00000000094778", n)))

    # ── ATC - 23012 ───────────────────────────────────────────────────────────
    n = "ATC - 23012"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 25, vz(aux("00000000086027", n)))
    esc(n, 45, vz(aux("00000000094775", n)))
    esc(n, 46, vz(aux("00000000094776", n)))
    esc(n, 47, vz(aux("00000000094777", n)))

    # ── CDP - 24790 ───────────────────────────────────────────────────────────
    n = "CDP - 24790"
    esc(n, 21,  vz(vds("00000000090819", n)))
    esc(n, 22,  vz(vds("00000000090821", n)))
    esc(n, 23,  vz(vds("00000000090822", n)))
    esc(n, 26,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 27,  vz(vds("00000000030190", n)))
    esc(n, 28,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)))
    esc(n, 31,  vz(vds("00000000089402", n)) + vz(vds("00000000003682", n)))
    esc(n, 32,  vz(vds("00000000089401", n)))
    esc(n, 33,  vz(vds("00000000003704", n)))
    esc(n, 39,  vz(vds("00000000054572", n)))
    esc(n, 42,  vz(vds("00000000046885", n)))
    esc(n, 43,  vz(vds("00000000046886", n)))
    esc(n, 44,  vz(vds("00000000046887", n)))
    esc(n, 48,  vz(vds("00000000046888", n)))
    esc(n, 49,  vz(vds("00000000046890", n)))
    esc(n, 50,  vz(vds("00000000046891", n)))
    esc(n, 64,  vz(vds("00000000052276", n)))
    esc(n, 65,  vz(vds("00000000052288", n)))
    esc(n, 66,  vz(vds("00000000053653", n)))
    esc(n, 67,  vz(vds("00000000053654", n)))
    esc(n, 68,  vz(vds("00000000053655", n)))
    esc(n, 69,  vz(vds("00000000053656", n)))
    esc(n, 87,  vz(vds("00000000085706", n)))
    esc(n, 88,  vz(vds("00000000094959", n)))
    esc(n, 89,  vz(vds("00000000095150", n)))
    esc(n, 90,  vz(vds("00000000085705", n)))
    esc(n, 91,  vz(vds("00000000085704", n)))
    esc(n, 107, vz(vds("00000000094371", n)))
    esc(n, 108, vz(vds("00000000095162", n)))
    esc(n, 109, vz(vds("00000000088796", n)))
    esc(n, 110, vz(vds("00000000088799", n)))
    esc(n, 116, vz(vds("00000000096442", n)))
    esc(n, 117, vz(vds("00000000096443", n)))
    esc(n, 119, vz(vds("00000000097241", n)))
    esc(n, 120, vz(vds("00000000096437", n)))
    esc(n, 121, vz(vds("00000000096438", n)))
    esc(n, 122, vz(vds("00000000096441", n)))

    # ── ER BOQ - 23614 ────────────────────────────────────────────────────────
    n = "ER BOQ - 23614"
    esc(n, 22,  vz(vds("00000000090819", n)))
    esc(n, 23,  vz(vds("00000000090821", n)))
    esc(n, 24,  vz(vds("00000000090822", n)))
    esc(n, 27,  vz(vds("00000000003591", n)))
    esc(n, 28,  vz(vds("00000000003608", n)))
    esc(n, 29,  vz(vds("00000000003610", n)))
    esc(n, 32,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 33,  vz(vds("00000000030190", n)))
    esc(n, 34,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)) + vz(vds("00000000089402", n)))
    esc(n, 37,  vz(vds("00000000003682", n)) + vz(vds("00000000089401", n)))
    esc(n, 38,  vz(vds("00000000003704", n)))
    esc(n, 39,  vz(vds("00000000054572", n)))
    esc(n, 46,  vz(vds("00000000046885", n)))
    esc(n, 47,  vz(vds("00000000046886", n)))
    esc(n, 48,  vz(vds("00000000046887", n)))
    esc(n, 55,  vz(vds("00000000046888", n)))
    esc(n, 56,  vz(vds("00000000046890", n)))
    esc(n, 57,  vz(vds("00000000046891", n)))
    esc(n, 78,  vz(vds("00000000052276", n)))
    esc(n, 79,  vz(vds("00000000052288", n)))
    esc(n, 80,  vz(vds("00000000053653", n)))
    esc(n, 81,  vz(vds("00000000053654", n)))
    esc(n, 82,  vz(vds("00000000053655", n)))
    esc(n, 83,  vz(vds("00000000053656", n)))
    esc(n, 88,  vz(vds("00000000090451", n)))
    esc(n, 93,  vz(vds("00000000087694", n)))
    esc(n, 96,  vz(vds("00000000085706", n)))
    esc(n, 97,  vz(vds("00000000094959", n)))
    esc(n, 98,  vz(vds("00000000095150", n)))
    esc(n, 99,  vz(vds("00000000085705", n)))
    esc(n, 100, vz(vds("00000000085704", n)))
    esc(n, 116, vz(vds("00000000094371", n)))
    esc(n, 117, vz(vds("00000000095162", n)))
    esc(n, 118, vz(vds("00000000088796", n)))
    esc(n, 119, vz(vds("00000000088799", n)))
    esc(n, 125, vz(vds("00000000096442", n)))
    esc(n, 126, vz(vds("00000000096443", n)))
    esc(n, 127, vz(vds("00000000096444", n)))
    esc(n, 128, vz(vds("00000000097241", n)))
    esc(n, 130, vz(vds("00000000096438", n)))
    esc(n, 131, vz(vds("00000000096441", n)))

    # ── ER PBE - 23343 ────────────────────────────────────────────────────────
    n = "ER PBE - 23343"
    esc(n, 22,  vz(vds("00000000090819", n)))
    esc(n, 23,  vz(vds("00000000090821", n)))
    esc(n, 24,  vz(vds("00000000090822", n)))
    esc(n, 27,  vz(vds("00000000003591", n)))
    esc(n, 28,  vz(vds("00000000003608", n)))
    esc(n, 29,  vz(vds("00000000003610", n)))
    esc(n, 32,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 33,  vz(vds("00000000030190", n)))
    esc(n, 34,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)) + vz(vds("00000000089402", n)))
    esc(n, 37,  vz(vds("00000000003682", n)) + vz(vds("00000000089401", n)))
    esc(n, 38,  vz(vds("00000000003704", n)))
    esc(n, 39,  vz(vds("00000000054572", n)))
    esc(n, 46,  vz(vds("00000000046885", n)))
    esc(n, 47,  vz(vds("00000000046886", n)))
    esc(n, 48,  vz(vds("00000000046887", n)))
    esc(n, 55,  vz(vds("00000000046888", n)))
    esc(n, 56,  vz(vds("00000000046890", n)))
    esc(n, 57,  vz(vds("00000000046891", n)))
    esc(n, 67,  vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)))
    esc(n, 68,  vz(vds("00000000056378", n)))  # 30 antes de 29 — invertido propositalmente
    esc(n, 69,  vz(vds("00000000056376", n)))
    esc(n, 78,  vz(vds("00000000052276", n)))
    esc(n, 79,  vz(vds("00000000052288", n)))
    esc(n, 80,  vz(vds("00000000053653", n)))
    esc(n, 81,  vz(vds("00000000053654", n)))
    esc(n, 82,  vz(vds("00000000053655", n)))
    esc(n, 83,  vz(vds("00000000053656", n)))
    esc(n, 88,  vz(vds("00000000090451", n)))
    esc(n, 96,  vz(vds("00000000085706", n)))
    esc(n, 99,  vz(vds("00000000085705", n)))
    esc(n, 116, vz(vds("00000000094371", n)))
    esc(n, 117, vz(vds("00000000095162", n)))
    esc(n, 118, vz(vds("00000000088796", n)))
    esc(n, 119, vz(vds("00000000088799", n)))
    esc(n, 125, vz(vds("00000000096442", n)))
    esc(n, 126, vz(vds("00000000096443", n)))
    esc(n, 128, vz(vds("00000000096444", n)))
    esc(n, 129, vz(vds("00000000096437", n)))  # linha 47 da VD ignorada
    esc(n, 130, vz(vds("00000000096438", n)))
    esc(n, 131, vz(vds("00000000096441", n)))

    # ── ER MG - 24119 ─────────────────────────────────────────────────────────
    n = "ER MG - 24119"
    esc(n, 22,  vz(vds("00000000090819", n)))
    esc(n, 23,  vz(vds("00000000090821", n)))
    esc(n, 24,  vz(vds("00000000090822", n)))
    esc(n, 27,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 28,  vz(vds("00000000030190", n)))
    esc(n, 29,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)) + vz(vds("00000000089402", n)))
    esc(n, 32,  vz(vds("00000000003682", n)) + vz(vds("00000000089401", n)))
    esc(n, 33,  vz(vds("00000000003704", n)))
    esc(n, 38,  vz(vds("00000000054572", n)))
    esc(n, 41,  vz(vds("00000000046885", n)))
    esc(n, 42,  vz(vds("00000000046886", n)))
    esc(n, 43,  vz(vds("00000000046887", n)))
    esc(n, 50,  vz(vds("00000000046888", n)))
    esc(n, 51,  vz(vds("00000000046890", n)))
    esc(n, 52,  vz(vds("00000000046891", n)))
    esc(n, 62,  vz(vds("00000000050110", n)) + vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)))
    esc(n, 63,  vz(vds("00000000056378", n)))  # 30 antes de 29 — invertido propositalmente
    esc(n, 64,  vz(vds("00000000056376", n)))
    esc(n, 73,  vz(vds("00000000052276", n)))
    esc(n, 74,  vz(vds("00000000052288", n)))
    esc(n, 75,  vz(vds("00000000053653", n)))
    esc(n, 76,  vz(vds("00000000053654", n)))
    esc(n, 77,  vz(vds("00000000053655", n)))
    esc(n, 78,  vz(vds("00000000053656", n)))
    esc(n, 91,  vz(vds("00000000085706", n)))
    esc(n, 94,  vz(vds("00000000085705", n)))
    esc(n, 95,  vz(vds("00000000085704", n)))
    esc(n, 98,  vz(vds("00000000094783", n)))
    esc(n, 99,  vz(vds("00000000094781", n)))
    esc(n, 102, vz(vds("00000000094775", n)))
    esc(n, 103, vz(vds("00000000094776", n)))
    esc(n, 104, vz(vds("00000000094778", n)))
    esc(n, 105, vz(vds("00000000094779", n)))
    esc(n, 111, vz(vds("00000000095162", n)))
    esc(n, 112, vz(vds("00000000088796", n)))
    esc(n, 113, vz(vds("00000000088799", n)))
    esc(n, 114, vz(vds("00000000096442", n)))
    esc(n, 120, vz(vds("00000000096443", n)))
    esc(n, 121, vz(vds("00000000096444", n)))
    esc(n, 123, vz(vds("00000000097241", n)))
    esc(n, 124, vz(vds("00000000096437", n)))
    esc(n, 125, vz(vds("00000000096438", n)))
    esc(n, 126, vz(vds("00000000096441", n)))

    return contagem


# ─────────────────────────────────────────────────────────────────────────────
# PREENCHIMENTO — JAÚ
# ─────────────────────────────────────────────────────────────────────────────

def preencher_jau(
    wb,
    leitor_aux: "LeitorAuxiliar",
    leitor_vds: "LeitorAuxiliar",
    col_novas: Dict[str, Dict[int, int]],
    dry_run: bool,
    logger: logging.Logger,
) -> Dict[str, int]:
    """
    Escreve valores nas abas da Planilha Jaú lendo de AUXILIAR e AUXILIAR VD'S.
    Acesso via código de produto + código de loja (imune a inserção/reordenação
    de linhas/colunas nas auxiliares).
    Retorna {nome_aba: quantidade_de_valores_escritos}.
    """
    contagem: Dict[str, int] = {}
    _warned: set = set()

    # Acessores curtos para legibilidade
    def aux(cod_prod: str, nome_aba: str):
        return leitor_aux.get(cod_prod, nome_aba)

    def vds(cod_prod: str, nome_aba: str):
        return leitor_vds.get(cod_prod, nome_aba)

    def esc(nome_aba: str, linha: int, valor: float, usou: int = 1) -> None:
        """Escreve `valor` em (linha, coluna USOU[usou]) da aba.
        usou=1 (default) → primeiro USOU (esquerda); usou=2 → segundo USOU; etc.
        """
        if nome_aba not in col_novas:
            if nome_aba not in _warned:
                if nome_aba in wb.sheetnames:
                    logger.warning("     '%s' sem coluna USOU — não preenchida", nome_aba)
                else:
                    logger.warning("     '%s' não existe na planilha — ignorada", nome_aba)
                _warned.add(nome_aba)
            return
        cols_usou = col_novas[nome_aba]
        if usou not in cols_usou:
            # Loja não tem esse USOU — silenciosamente ignora (não é erro)
            return
        if not dry_run:
            wb[nome_aba].cell(row=linha, column=cols_usou[usou]).value = valor
        contagem[nome_aba] = contagem.get(nome_aba, 0) + 1

    # ── JC 3822 ───────────────────────────────────────────────────────────────
    n = "JC 3822"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 29, vz(aux("00000000090774", n)))
    esc(n, 30, vz(aux("00000000090770", n)))
    esc(n, 31, vz(aux("00000000090772", n)))
    esc(n, 32, vz(aux("00000000094682", n)))
    esc(n, 33, vz(aux("00000000094775", n)))
    esc(n, 34, vz(aux("00000000094776", n)))
    esc(n, 37, vz(aux("00000000095151", n)))

    # ── JD 14446 ──────────────────────────────────────────────────────────────
    n = "JD 14446"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 34, vz(aux("00000000095151", n)))

    # ── BB 12066 ──────────────────────────────────────────────────────────────
    n = "BB 12066"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 29, vz(aux("00000000094775", n)))
    esc(n, 30, vz(aux("00000000094776", n)))
    esc(n, 33, vz(aux("00000000095151", n)))
    esc(n, 34, vz(aux("00000000094682", n)))

    # ── SM 23048 ──────────────────────────────────────────────────────────────
    n = "SM 23048"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 34, vz(aux("00000000095151", n)))

    # ── JSH 11722 ─────────────────────────────────────────────────────────────
    n = "JSH 11722"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 27, vz(aux("00000000094802", n)))
    esc(n, 28, vz(aux("00000000094775", n)))
    esc(n, 29, vz(aux("00000000094776", n)))
    esc(n, 32, vz(aux("00000000095151", n)))
    esc(n, 33, vz(aux("00000000094682", n)))

    # ── CONF 14553 ────────────────────────────────────────────────────────────
    n = "CONF 14553"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 33, vz(aux("00000000094682", n)))

    # ── DC 7529 ───────────────────────────────────────────────────────────────
    n = "DC 7529"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 46, vz(aux("00000000095151", n)))

    # ── IT 6942 ───────────────────────────────────────────────────────────────
    n = "IT 6942"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 27, vz(aux("00000000087694", n)))
    esc(n, 28, vz(aux("00000000094802", n)))
    esc(n, 29, vz(aux("00000000094775", n)))
    esc(n, 30, vz(aux("00000000094776", n)))
    esc(n, 34, vz(aux("00000000094682", n)))

    # ── BR 6954 ───────────────────────────────────────────────────────────────
    n = "BR 6954"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 46, vz(aux("00000000094777", n)))

    # ── CDJ 23091 ─────────────────────────────────────────────────────────────
    n = "CDJ 23091"
    esc(n, 14,  vz(vds("00000000090819", n)))
    esc(n, 15,  vz(vds("00000000090821", n)))
    esc(n, 16,  vz(vds("00000000090822", n)))
    esc(n, 19,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 20,  vz(vds("00000000030190", n)))
    esc(n, 21,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)))
    esc(n, 24,  vz(vds("00000000089402", n)))
    esc(n, 25,  vz(vds("00000000089401", n)))
    esc(n, 26,  vz(vds("00000000003704", n)))
    esc(n, 30,  vz(vds("00000000054572", n)))
    esc(n, 33,  vz(vds("00000000046885", n)))
    esc(n, 34,  vz(vds("00000000046886", n)))
    esc(n, 35,  vz(vds("00000000046887", n)))
    esc(n, 39,  vz(vds("00000000046888", n)))
    esc(n, 40,  vz(vds("00000000046890", n)))
    esc(n, 41,  vz(vds("00000000046891", n)))
    esc(n, 47,  vz(vds("00000000050110", n)) + vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)))
    esc(n, 48,  vz(vds("00000000056378", n)))
    esc(n, 49,  vz(vds("00000000056376", n)))
    esc(n, 55,  vz(vds("00000000052276", n)))
    esc(n, 56,  vz(vds("00000000052288", n)))
    esc(n, 57,  vz(vds("00000000053653", n)))
    esc(n, 58,  vz(vds("00000000053654", n)))
    esc(n, 59,  vz(vds("00000000053655", n)))
    esc(n, 60,  vz(vds("00000000053656", n)))
    esc(n, 80,  vz(vds("00000000095150", n)))
    esc(n, 81,  vz(vds("00000000085705", n)))
    esc(n, 82,  vz(vds("00000000085704", n)))
    esc(n, 89,  vz(vds("00000000094775", n)))
    esc(n, 90,  vz(vds("00000000094776", n)))
    esc(n, 91,  vz(vds("00000000094778", n)))
    esc(n, 92,  vz(vds("00000000094779", n)))
    esc(n, 104, vz(vds("00000000094371", n)))
    esc(n, 105, vz(vds("00000000095162", n)))
    esc(n, 106, vz(vds("00000000088796", n)))
    esc(n, 107, vz(vds("00000000088799", n)))
    esc(n, 113, vz(vds("00000000096443", n)))
    esc(n, 114, vz(vds("00000000096444", n)))
    esc(n, 115, vz(vds("00000000097241", n)))
    esc(n, 116, vz(vds("00000000096437", n)))
    esc(n, 117, vz(vds("00000000096438", n)))
    esc(n, 118, vz(vds("00000000096441", n)))

    # ── ERJ 22838 ─────────────────────────────────────────────────────────────
    n = "ERJ 22838"
    esc(n, 22,  vz(vds("00000000003591", n)))
    esc(n, 23,  vz(vds("00000000003608", n)))
    esc(n, 24,  vz(vds("00000000003610", n)))
    esc(n, 27,  vz(vds("00000000090819", n)))
    esc(n, 28,  vz(vds("00000000090821", n)))
    esc(n, 29,  vz(vds("00000000090822", n)))
    esc(n, 32,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 33,  vz(vds("00000000030190", n)))
    esc(n, 34,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)))
    esc(n, 37,  vz(vds("00000000089402", n)) + vz(vds("00000000003682", n)))
    esc(n, 38,  vz(vds("00000000089401", n)))
    esc(n, 39,  vz(vds("00000000003704", n)))
    esc(n, 43,  vz(vds("00000000054572", n)))
    esc(n, 46,  vz(vds("00000000046885", n)))
    esc(n, 47,  vz(vds("00000000046886", n)))
    esc(n, 48,  vz(vds("00000000046887", n)))
    esc(n, 55,  vz(vds("00000000046888", n)))
    esc(n, 56,  vz(vds("00000000046890", n)))
    esc(n, 57,  vz(vds("00000000046891", n)))
    esc(n, 67,  vz(vds("00000000050110", n)) + vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)))
    esc(n, 68,  vz(vds("00000000056378", n)))
    esc(n, 69,  vz(vds("00000000056376", n)))
    esc(n, 78,  vz(vds("00000000052276", n)))
    esc(n, 79,  vz(vds("00000000052288", n)))
    esc(n, 80,  vz(vds("00000000053653", n)))
    esc(n, 81,  vz(vds("00000000053654", n)))
    esc(n, 82,  vz(vds("00000000053655", n)))
    esc(n, 83,  vz(vds("00000000053656", n)))
    esc(n, 88,  vz(vds("00000000090451", n)))
    esc(n, 92,  vz(vds("00000000089713", n)))
    esc(n, 93,  vz(vds("00000000087694", n)))
    esc(n, 96,  vz(vds("00000000085706", n)))
    esc(n, 97,  vz(vds("00000000094959", n)))
    esc(n, 98,  vz(vds("00000000095150", n)))
    esc(n, 99,  vz(vds("00000000085705", n)))
    esc(n, 100, vz(vds("00000000085704", n)))
    esc(n, 107, vz(vds("00000000094775", n)))
    esc(n, 108, vz(vds("00000000094776", n)))
    esc(n, 109, vz(vds("00000000094778", n)))
    esc(n, 110, vz(vds("00000000094779", n)))
    esc(n, 118, vz(vds("00000000094371", n)))
    esc(n, 119, vz(vds("00000000095162", n)))
    esc(n, 120, vz(vds("00000000088796", n)))
    esc(n, 121, vz(vds("00000000088799", n)))
    esc(n, 127, vz(vds("00000000096442", n)))
    esc(n, 128, vz(vds("00000000096444", n)))
    esc(n, 129, vz(vds("00000000097241", n)))
    esc(n, 130, vz(vds("00000000096437", n)))
    esc(n, 131, vz(vds("00000000096438", n)))
    esc(n, 132, vz(vds("00000000096441", n)))

    # ── ER SM 24137 ───────────────────────────────────────────────────────────
    n = "ER SM 24137"
    esc(n, 22,  vz(vds("00000000090819", n)))
    esc(n, 23,  vz(vds("00000000090821", n)))
    esc(n, 24,  vz(vds("00000000090822", n)))
    esc(n, 27,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 28,  vz(vds("00000000030190", n)))
    esc(n, 29,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)) + vz(vds("00000000089402", n)))
    esc(n, 32,  vz(vds("00000000003682", n)) + vz(vds("00000000089401", n)))
    esc(n, 33,  vz(vds("00000000089401", n)))  # VD row 16 usado novamente — replicado do VBA
    esc(n, 34,  vz(vds("00000000003704", n)))
    esc(n, 41,  vz(vds("00000000046885", n)))
    esc(n, 42,  vz(vds("00000000046886", n)))
    esc(n, 43,  vz(vds("00000000046887", n)))
    esc(n, 50,  vz(vds("00000000046888", n)))
    esc(n, 51,  vz(vds("00000000046890", n)))
    esc(n, 52,  vz(vds("00000000046891", n)))
    esc(n, 62,  vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)) + vz(vds("00000000056376", n)))
    esc(n, 63,  vz(vds("00000000056378", n)))
    esc(n, 64,  vz(vds("00000000056376", n)))  # VD row 29 usado novamente — replicado do VBA
    esc(n, 73,  vz(vds("00000000052276", n)))
    esc(n, 74,  vz(vds("00000000052288", n)))
    esc(n, 75,  vz(vds("00000000053653", n)))
    esc(n, 76,  vz(vds("00000000053654", n)))
    esc(n, 77,  vz(vds("00000000053655", n)))
    esc(n, 78,  vz(vds("00000000053656", n)))
    esc(n, 81,  vz(vds("00000000085706", n)))
    esc(n, 82,  vz(vds("00000000094959", n)))
    esc(n, 83,  vz(vds("00000000095150", n)))
    esc(n, 84,  vz(vds("00000000085705", n)))
    esc(n, 85,  vz(vds("00000000085704", n)))
    esc(n, 92,  vz(vds("00000000094775", n)))
    esc(n, 93,  vz(vds("00000000094776", n)))
    esc(n, 94,  vz(vds("00000000094778", n)))
    esc(n, 95,  vz(vds("00000000094779", n)))
    esc(n, 100, vz(vds("00000000090451", n)))
    esc(n, 108, vz(vds("00000000094371", n)))
    esc(n, 109, vz(vds("00000000095162", n)))
    esc(n, 110, vz(vds("00000000088796", n)))
    esc(n, 111, vz(vds("00000000088799", n)))
    esc(n, 117, vz(vds("00000000096443", n)))
    esc(n, 118, vz(vds("00000000096444", n)))
    esc(n, 119, vz(vds("00000000097241", n)))
    esc(n, 120, vz(vds("00000000096437", n)))
    esc(n, 121, vz(vds("00000000096438", n)))
    esc(n, 122, vz(vds("00000000096441", n)))
    esc(n, 125, vz(vds("00000000087694", n)))

    return contagem


# ─────────────────────────────────────────────────────────────────────────────
# PREENCHIMENTO — BAURU
# ─────────────────────────────────────────────────────────────────────────────

def preencher_bauru(
    wb,
    leitor_aux: "LeitorAuxiliar",
    leitor_vds: "LeitorAuxiliar",
    col_novas: Dict[str, Dict[int, int]],
    dry_run: bool,
    logger: logging.Logger,
) -> Dict[str, int]:
    """
    Escreve valores nas abas da Planilha Bauru lendo de AUXILIAR e AUXILIAR VD'S.
    Acesso via código de produto + código de loja (imune a inserção/reordenação
    de linhas/colunas nas auxiliares).
    Retorna {nome_aba: quantidade_de_valores_escritos}.
    """
    contagem: Dict[str, int] = {}
    _warned: set = set()

    # Acessores curtos para legibilidade
    def aux(cod_prod: str, nome_aba: str):
        return leitor_aux.get(cod_prod, nome_aba)

    def vds(cod_prod: str, nome_aba: str):
        return leitor_vds.get(cod_prod, nome_aba)

    def esc(nome_aba: str, linha: int, valor: float, usou: int = 1) -> None:
        """Escreve `valor` em (linha, coluna USOU[usou]) da aba.
        usou=1 (default) → primeiro USOU (esquerda); usou=2 → segundo USOU; etc.
        """
        if nome_aba not in col_novas:
            if nome_aba not in _warned:
                if nome_aba in wb.sheetnames:
                    logger.warning("     '%s' sem coluna USOU — não preenchida", nome_aba)
                else:
                    logger.warning("    '%s' não existe na planilha — ignorada", nome_aba)
                _warned.add(nome_aba)
            return
        cols_usou = col_novas[nome_aba]
        if usou not in cols_usou:
            # Loja não tem esse USOU — silenciosamente ignora (não é erro)
            return
        if not dry_run:
            wb[nome_aba].cell(row=linha, column=cols_usou[usou]).value = valor
        contagem[nome_aba] = contagem.get(nome_aba, 0) + 1

    # ── BSH 6700 ──────────────────────────────────────────────────────────────
    n = "BSH 6700"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))

    # ── BOUL 13868 ────────────────────────────────────────────────────────────
    n = "BOUL 13868"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 48, vz(aux("00000000094776", n)))

    # ── TT 13370 ──────────────────────────────────────────────────────────────
    n = "TT 13370"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 46, vz(aux("00000000094775", n)))
    esc(n, 47, vz(aux("00000000094776", n)))

    # ── TDQ 20000 ─────────────────────────────────────────────────────────────
    n = "TDQ 20000"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 47, vz(aux("00000000094779", n)))

    # ── GET 23049 ─────────────────────────────────────────────────────────────
    n = "GET 23049"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 26, vz(aux("00000000086027", n)))
    esc(n, 32, vz(aux("00000000087691", n)))
    esc(n, 33, vz(aux("00000000087694", n)))
    esc(n, 46, vz(aux("00000000094778", n)))
    esc(n, 47, vz(aux("00000000094779", n)))

    # ── Q7 6727 ───────────────────────────────────────────────────────────────
    n = "Q7 6727"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 33, vz(aux("00000000087691", n)))
    esc(n, 34, vz(aux("00000000087694", n)))
    esc(n, 44, vz(aux("00000000094775", n)))
    esc(n, 48, vz(aux("00000000095151", n)))

    # ── Q2 12466 ──────────────────────────────────────────────────────────────
    n = "Q2 12466"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 33, vz(aux("00000000087691", n)))
    esc(n, 34, vz(aux("00000000087694", n)))
    esc(n, 47, vz(aux("00000000094779", n)))
    esc(n, 48, vz(aux("00000000095151", n)))

    # ── MD 12942 ──────────────────────────────────────────────────────────────
    n = "MD 12942"
    esc(n, 20, vz(aux("00000000089402", n)) + vz(aux("00000000003517", n)))
    esc(n, 21, vz(aux("00000000089401", n)) + vz(aux("00000000003530", n)))
    esc(n, 22, vz(aux("00000000003554", n)))
    esc(n, 23, vz(aux("00000000089400", n)))
    esc(n, 27, vz(aux("00000000086027", n)))
    esc(n, 33, vz(aux("00000000087691", n)))
    esc(n, 47, vz(aux("00000000094776", n)))

    # ── CDB - 23280 ───────────────────────────────────────────────────────────
    n = "CDB - 23280"
    esc(n, 16,  vz(vds("00000000090819", n)))
    esc(n, 17,  vz(vds("00000000090821", n)))
    esc(n, 18,  vz(vds("00000000090822", n)))
    esc(n, 21,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 22,  vz(vds("00000000030190", n)))
    esc(n, 23,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)))
    esc(n, 25,  vz(vds("00000000003704", n)))  # mesmo valor da linha 28 — replicado do VBA
    esc(n, 26,  vz(vds("00000000089402", n)))
    esc(n, 27,  vz(vds("00000000089401", n)))
    esc(n, 28,  vz(vds("00000000003704", n)))
    esc(n, 33,  vz(vds("00000000054572", n)))
    esc(n, 36,  vz(vds("00000000046885", n)))
    esc(n, 37,  vz(vds("00000000046886", n)))
    esc(n, 38,  vz(vds("00000000046887", n)))
    esc(n, 42,  vz(vds("00000000046888", n)))
    esc(n, 43,  vz(vds("00000000046890", n)))
    esc(n, 44,  vz(vds("00000000046891", n)))
    esc(n, 50,  vz(vds("00000000050110", n)) + vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)))
    esc(n, 51,  vz(vds("00000000056378", n)))
    esc(n, 52,  vz(vds("00000000056376", n)))
    esc(n, 58,  vz(vds("00000000052276", n)))
    esc(n, 59,  vz(vds("00000000052288", n)))
    esc(n, 60,  vz(vds("00000000053653", n)))
    esc(n, 61,  vz(vds("00000000053654", n)))
    esc(n, 62,  vz(vds("00000000053655", n)))
    esc(n, 63,  vz(vds("00000000053656", n)))
    esc(n, 73,  vz(vds("00000000090451", n)))
    esc(n, 76,  vz(vds("00000000089713", n)))
    esc(n, 78,  vz(vds("00000000087694", n)))
    esc(n, 81,  vz(vds("00000000085706", n)))
    esc(n, 82,  vz(vds("00000000094959", n)))
    esc(n, 83,  vz(vds("00000000095150", n)))
    esc(n, 84,  vz(vds("00000000085705", n)))
    esc(n, 85,  vz(vds("00000000085704", n)))
    esc(n, 88,  vz(vds("00000000094783", n)))
    esc(n, 89,  vz(vds("00000000094781", n)))
    esc(n, 92,  vz(vds("00000000094775", n)))
    esc(n, 93,  vz(vds("00000000094776", n)))
    esc(n, 94,  vz(vds("00000000094778", n)))
    esc(n, 95,  vz(vds("00000000094779", n)))
    esc(n, 107, vz(vds("00000000094371", n)))
    esc(n, 108, vz(vds("00000000095162", n)))
    esc(n, 109, vz(vds("00000000088796", n)))
    esc(n, 110, vz(vds("00000000088799", n)))
    esc(n, 116, vz(vds("00000000096442", n)))
    esc(n, 117, vz(vds("00000000096443", n)))
    esc(n, 118, vz(vds("00000000097241", n)))
    esc(n, 119, vz(vds("00000000096437", n)))
    esc(n, 120, vz(vds("00000000096438", n)))
    esc(n, 121, vz(vds("00000000096441", n)))

    # ── ERB - 22851 ───────────────────────────────────────────────────────────
    n = "ERB - 22851"
    esc(n, 22,  vz(vds("00000000090819", n)))
    esc(n, 23,  vz(vds("00000000090821", n)))
    esc(n, 24,  vz(vds("00000000090822", n)))
    esc(n, 27,  vz(vds("00000000094389", n)) + vz(vds("00000000043673", n)))
    esc(n, 28,  vz(vds("00000000030190", n)))
    esc(n, 29,  vz(vds("00000000045033", n)) + vz(vds("00000000043763", n)) + vz(vds("00000000030194", n)))
    esc(n, 32,  vz(vds("00000000089402", n)))
    esc(n, 33,  vz(vds("00000000089401", n)))
    esc(n, 34,  vz(vds("00000000003704", n)))
    esc(n, 41,  vz(vds("00000000046885", n)))
    esc(n, 42,  vz(vds("00000000046886", n)))
    esc(n, 43,  vz(vds("00000000046887", n)))
    esc(n, 50,  vz(vds("00000000046888", n)))
    esc(n, 51,  vz(vds("00000000046890", n)))
    esc(n, 52,  vz(vds("00000000046891", n)))
    esc(n, 62,  vz(vds("00000000050110", n)) + vz(vds("00000000056369", n)) + vz(vds("00000000056375", n)))
    esc(n, 63,  vz(vds("00000000056378", n)))
    esc(n, 64,  vz(vds("00000000056376", n)))
    esc(n, 73,  vz(vds("00000000052276", n)))
    esc(n, 74,  vz(vds("00000000052288", n)))
    esc(n, 75,  vz(vds("00000000053653", n)))
    esc(n, 76,  vz(vds("00000000053654", n)))
    esc(n, 77,  vz(vds("00000000053655", n)))
    esc(n, 78,  vz(vds("00000000053656", n)))
    esc(n, 83,  vz(vds("00000000090451", n)))
    esc(n, 88,  vz(vds("00000000087694", n)))
    esc(n, 93,  vz(vds("00000000095150", n)))
    esc(n, 94,  vz(vds("00000000085705", n)))
    esc(n, 95,  vz(vds("00000000085704", n)))
    esc(n, 102, vz(vds("00000000094775", n)))
    esc(n, 117, vz(vds("00000000094371", n)))
    esc(n, 118, vz(vds("00000000095162", n)))
    esc(n, 119, vz(vds("00000000088796", n)))
    esc(n, 120, vz(vds("00000000088799", n)))
    esc(n, 126, vz(vds("00000000096442", n)))
    esc(n, 127, vz(vds("00000000096443", n)))
    esc(n, 128, vz(vds("00000000096444", n)))
    esc(n, 129, vz(vds("00000000096437", n)))
    esc(n, 130, vz(vds("00000000096438", n)))
    esc(n, 131, vz(vds("00000000096441", n)))

    return contagem


# ─────────────────────────────────────────────────────────────────────────────
# INSERÇÃO DE COLUNA — SEÇÃO "NOVO PEDIDO"
# ─────────────────────────────────────────────────────────────────────────────

def inserir_coluna_novo_pedido(
    ws, dry_run: bool, logger: logging.Logger, nome_aba: str
) -> Optional[int]:
    """
    Localiza a seção 'Novo Pedido', insere uma nova coluna em col_ancora+1 e
    escreve data de hoje + zeros, deslocando as datas anteriores para a direita.
    Mantém 3 colunas visíveis (nova + 2 anteriores), oculta o restante.
    """
    # ── 1. Encontrar coluna âncora com texto 'NOVO PEDIDO' ───────────────────
    col_ancora = None
    for col in range(1, ws.max_column + 1):
        for row in range(1, min(ws.max_row + 1, 201)):
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().upper() == "NOVO PEDIDO":
                col_ancora = col
                break
        if col_ancora is not None:
            break

    if col_ancora is None:
        return None

    # ── 2. Bloco de datas existentes a partir de col_ancora + 1 (antes da inserção)
    primeira_data = None
    for col in range(col_ancora + 1, min(col_ancora + 32, ws.max_column + 1)):
        if any(_cel_tem_data(ws.cell(row=row, column=col)) for row in range(1, 16)):
            primeira_data = col
            break

    if primeira_data is None:
        return None

    date_cols_antes: list = []
    for col in range(primeira_data, ws.max_column + 1):
        if any(_cel_tem_data(ws.cell(row=row, column=col)) for row in range(1, 16)):
            date_cols_antes.append(col)
        else:
            break

    template_col_antes = date_cols_antes[0]

    # Inserir imediatamente antes da primeira data existente (sem pular colunas em branco)
    col_alvo = primeira_data

    if dry_run:
        logger.debug(
            "    [DRY-RUN] Aba '%s': Novo Pedido — inserir em col %d, template col %d (%d datas)",
            nome_aba, col_alvo, template_col_antes, len(date_cols_antes),
        )
        return col_alvo

    ultima_linha = ws.max_row

    # ── 4. Snapshot de larguras antes da inserção ─────────────────────────────
    all_widths: dict = {}
    all_hidden: dict = {}
    for letra, dim in ws.column_dimensions.items():
        col_num = column_index_from_string(letra)
        if dim.width and dim.width > 0:
            all_widths[col_num] = dim.width
        if dim.hidden:
            all_hidden[col_num] = True

    # ── 5. Inserir nova coluna em col_alvo ────────────────────────────────────
    ws.insert_cols(col_alvo)

    # Após a inserção todos os índices >= col_alvo deslocaram +1
    template_col = template_col_antes + 1
    date_cols = [c + 1 for c in date_cols_antes]

    # ── 6. Copiar formato e escrever valores na nova coluna ───────────────────
    copiar_formato(ws, template_col, col_alvo)

    for row in range(1, ultima_linha + 1):
        tmpl = ws.cell(row=row, column=template_col)
        nova = ws.cell(row=row, column=col_alvo)
        if isinstance(nova, MergedCell):
            continue
        nova.value = None
        if _cel_tem_data(tmpl):
            nova.value = dt.datetime.today()
            nova.number_format = "DD/MM"
        elif isinstance(tmpl.value, (int, float)) and tmpl.value is not None:
            nova.value = 0

    # ── 6b. Corrigir fórmulas das colunas USOU deslocadas por esta inserção ─────
    # A inserção acima empurrou +1 todas as colunas USOU à direita de col_alvo.
    # As fórmulas que foram geradas em inserir_coluna ficaram com referências
    # uma coluna atrás do correto — aqui as reescrevemos na posição definitiva.
    usou_cols_apos = _encontrar_todas_col_usou(ws)
    for usou_col in usou_cols_apos:
        if usou_col <= col_alvo:
            continue  # USOU à esquerda do ponto de inserção: não foi deslocado
        if usou_col < 3:
            continue
        letra_prev2 = get_column_letter(usou_col - 2)
        letra_prev1 = get_column_letter(usou_col - 1)
        letra_self  = get_column_letter(usou_col)
        for row in range(1, ultima_linha + 1):
            cell = ws.cell(row=row, column=usou_col)
            if (cell.value is not None
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")):
                cell.value = (
                    f"={letra_prev2}{row}"
                    f"+IF(NOT(ISBLANK(OFFSET({letra_self}{row},0,6))),"
                    f"N(OFFSET({letra_self}{row},0,6)),"
                    f"N(OFFSET({letra_self}{row},0,7)))"
                    f"-{letra_prev1}{row}"
                )

    # ── 7. Reconstruir larguras com deslocamento de +1 para cols >= col_alvo ──
    ws.column_dimensions.clear()
    for orig_col, width in all_widths.items():
        nova_letra = get_column_letter(orig_col + 1 if orig_col >= col_alvo else orig_col)
        ws.column_dimensions[nova_letra].width = width
    for orig_col in all_hidden:
        nova_letra = get_column_letter(orig_col + 1 if orig_col >= col_alvo else orig_col)
        ws.column_dimensions[nova_letra].hidden = True
    # Nova coluna herda largura do template
    if template_col_antes in all_widths:
        ws.column_dimensions[get_column_letter(col_alvo)].width = all_widths[template_col_antes]

    # ── 8. Visibilidade: col_alvo + 2 datas anteriores = 3 visíveis ──────────
    ordered_cols = [col_alvo] + date_cols
    for i, col in enumerate(ordered_cols):
        ws.column_dimensions[get_column_letter(col)].hidden = (i >= 3)

    logger.debug(
        "    Aba '%s': Novo Pedido — inserida em col %d, %d semanas (%d ocultas)",
        nome_aba, col_alvo, len(ordered_cols), max(0, len(ordered_cols) - 3),
    )
    return col_alvo


# ─────────────────────────────────────────────────────────────────────────────
# AVANÇAR REFERÊNCIAS NAS ABAS DE LOJAS
# ─────────────────────────────────────────────────────────────────────────────

def _avancar_col_ref(formula: str) -> str:
    """
    Transforma ='NomeAba'!CF8 → ='NomeAba'!CG8 (avança a coluna +1).
    Retorna a fórmula inalterada se o padrão não for reconhecido.
    """
    sep = formula.find("'!")          # posição do fecho de aspas + exclamação
    if sep == -1:
        return formula
    prefixo = formula[:sep + 2]       # ex: ='MG 11733'!
    resto   = formula[sep + 2:]       # ex: CF8
    # Separar letras da coluna e número da linha
    i = 0
    while i < len(resto) and resto[i].isalpha():
        i += 1
    if i == 0 or not resto[i:].isdigit():
        return formula                # formato inesperado — não altera
    col_letra = resto[:i]
    linha     = resto[i:]
    nova_col  = get_column_letter(column_index_from_string(col_letra.upper()) + 1)
    return f"{prefixo}{nova_col}{linha}"


def avancar_referencias_lojas(
    ws, dry_run: bool, logger: logging.Logger, nome_aba: str
) -> int:
    """
    Percorre todas as células da aba de resumo (LOJAS PRAIA / LOJAS BAURU / LOJAS JAÚ)
    e avança +1 coluna em cada fórmula do tipo ='NomeAba'!COLlinha.
    Retorna o número de fórmulas alteradas.
    """
    total = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            val = cell.value
            # Só interessa fórmulas com referência externa: ='NomeAba'!...
            if not (isinstance(val, str) and val.startswith("='") and "'!" in val):
                continue
            nova = _avancar_col_ref(val)
            if nova == val:
                continue
            if not dry_run:
                cell.value = nova
            total += 1

    if dry_run:
        logger.debug(
            "    [DRY-RUN] Aba '%s': %d referência(s) seriam avançadas", nome_aba, total
        )
    else:
        logger.debug(
            "    Aba '%s': %d referência(s) avançadas", nome_aba, total
        )
    return total


# ─────────────────────────────────────────────────────────────────────────────
# PROCESSAR UMA PLANILHA PAI
# ─────────────────────────────────────────────────────────────────────────────

def processar_planilha(
    regiao: str,
    caminho_xlsm: Path,
    leitor_aux: "LeitorAuxiliar",
    leitor_vds: "LeitorAuxiliar",
    preencher_fn: Callable,
    dry_run: bool,
    logger: logging.Logger,
) -> None:
    """Abre o .xlsm, insere colunas em todas as abas elegíveis, preenche valores e salva."""
    label = regiao.upper()
    print(f"\n  [{label}] Processando {caminho_xlsm.name}")
    logger.info("[%s] Abrindo: %s", label, caminho_xlsm)

    if not caminho_xlsm.exists():
        msg = f"Arquivo não encontrado: {caminho_xlsm}"
        logger.error(msg)
        print(f"    {msg}")
        return

    try:
        wb = openpyxl.load_workbook(str(caminho_xlsm), keep_vba=True)
    except PermissionError:
        msg = f"{caminho_xlsm.name} está aberto em outro programa — feche e tente novamente"
        logger.error(msg)
        print(f"     ERRO: {msg}")
        return
    except Exception as exc:
        logger.error("Falha ao abrir %s: %s", caminho_xlsm.name, exc)
        print(f"     Falha ao abrir: {exc}")
        return

    # Para cada aba: {índice_usou: posição_da_nova_coluna}
    # Ex.: {' PL 12973': {1: 28, 2: 60}, 'BOQ 11734': {1: 79}}
    col_novas: Dict[str, Dict[int, int]] = {}
    abas_sem_usou: list = []

    for nome_aba in wb.sheetnames:
        if not eh_elegivel(nome_aba):
            continue
        ws = wb[nome_aba]
        # USOU primeiro — insere colunas e oculta semanas antigas
        # Novo Pedido depois — sua visibilidade sobrescreve o que o USOU tiver ocultado
        cols_dict = inserir_coluna(ws, dry_run, logger, nome_aba)
        if cols_dict is None:
            abas_sem_usou.append(nome_aba)
        else:
            col_novas[nome_aba] = cols_dict
        inserir_coluna_novo_pedido(ws, dry_run, logger, nome_aba)

    # Preencher valores das lojas
    contagem = preencher_fn(wb, leitor_aux, leitor_vds, col_novas, dry_run, logger)

    # Exibir resultado por loja
    for nome_aba, n_vals in contagem.items():
        cols_dict = col_novas.get(nome_aba)
        if cols_dict:
            col = ", ".join(f"USOU{k}={v}" for k, v in cols_dict.items())
        else:
            col = "?"
        print(f"     {nome_aba:<22} — {col}, {n_vals} valores escritos")

    if abas_sem_usou:
        nomes = ", ".join(abas_sem_usou)
        print(f"     {len(abas_sem_usou)} aba(s) sem coluna USOU (ignoradas): {nomes}")
        logger.warning("[%s] Abas sem USOU: %s", label, nomes)

    # Avançar referências nas abas de resumo de lojas (LOJAS PRAIA / LOJAS BAURU / LOJAS JAÚ)
    for nome_aba in wb.sheetnames:
        if nome_aba not in _ABAS_LOJAS:
            continue
        ws = wb[nome_aba]
        n = avancar_referencias_lojas(ws, dry_run, logger, nome_aba)
        if n > 0:
            print(f"     {nome_aba:<22} — {n} referência(s) avançadas")

    if not dry_run:
        try:
            wb.save(str(caminho_xlsm))
            logger.info("[%s] Salvo: %s", label, caminho_xlsm.name)
        except PermissionError:
            msg = f"{caminho_xlsm.name} está aberto em outro programa — não foi possível salvar"
            logger.error(msg)
            print(f"     ERRO ao salvar: {msg}")
        except Exception as exc:
            logger.error("Erro ao salvar %s: %s", caminho_xlsm.name, exc)
            print(f"     Erro ao salvar: {exc}")
    else:
        logger.info("[DRY-RUN] %s não foi modificado", caminho_xlsm.name)

    wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────

def _setup_logging(log_dir: Path, dry_run: bool) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    hoje = dt.date.today().strftime("%Y-%m-%d")
    sufixo = "_dry-run" if dry_run else ""
    log_path = log_dir / f"processo1_{hoje}{sufixo}.log"

    logger = logging.getLogger("processo1")
    logger.setLevel(logging.DEBUG)

    fh = logging.FileHandler(str(log_path), encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)

    # Só warnings e acima no console (os prints cuidam do output normal)
    ch = logging.StreamHandler(sys.stderr)
    ch.setLevel(logging.WARNING)
    ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(ch)

    return logger


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    # Garante UTF-8 no terminal Windows
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description="Processo 1 — Inserção de coluna e preenchimento das Planilhas Pai Antilhas"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Simula a execução sem escrever nem fazer backup")
    parser.add_argument("--regiao", choices=["praia", "jau", "bauru"],
                        help="Processar apenas uma região")
    args = parser.parse_args()

    script_dir = Path(__file__).parent

    # Carregar configuração
    config_path = script_dir / "config_p1.yaml"
    if not config_path.exists():
        sys.exit(f"ERRO: config_p1.yaml não encontrado em {script_dir}")
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    def resolve(p: str) -> Path:
        return (script_dir / p).resolve()

    logger = _setup_logging(resolve(cfg["dirs"]["logs"]), args.dry_run)

    # Cabeçalho
    hoje_fmt = dt.date.today().strftime("%d/%m/%Y")
    dry_tag = " [DRY-RUN]" if args.dry_run else ""
    print("=" * 60)
    print(f"  PROCESSO 1 — ANTILHAS — {hoje_fmt}{dry_tag}")
    print("=" * 60)
    logger.info("Iniciando Processo 1%s", dry_tag)

    # Backup antes de qualquer modificação
    backup_cfg = {
        "dirs": {"backup": str(resolve(cfg["dirs"]["backup"]))},
        "planilhas_pai": {k: str(resolve(v)) for k, v in cfg["planilhas_pai"].items()},
    }
    backup(backup_cfg, args.dry_run, logger)

    # Abrir AUXILIAR.xlsx
    path_aux = resolve(cfg["auxiliar"]["path"])
    try:
        wb_aux = openpyxl.load_workbook(str(path_aux), data_only=True)
        aba_aux = cfg["auxiliar"].get("aba") or wb_aux.sheetnames[0]
        wsAux = wb_aux[aba_aux]
    except FileNotFoundError:
        sys.exit(f"ERRO: AUXILIAR.xlsx não encontrado em {path_aux}")
    except Exception as exc:
        sys.exit(f"ERRO ao abrir AUXILIAR.xlsx: {exc}")

    # Abrir AUXILIAR_VDS.xlsx
    path_vds = resolve(cfg["auxiliar_vds"]["path"])
    try:
        wb_vds = openpyxl.load_workbook(str(path_vds), data_only=True)
        aba_vds = cfg["auxiliar_vds"].get("aba") or wb_vds.sheetnames[0]
        wsVD = wb_vds[aba_vds]
    except FileNotFoundError:
        sys.exit(f"ERRO: AUXILIAR_VDS.xlsx não encontrado em {path_vds}")
    except Exception as exc:
        sys.exit(f"ERRO ao abrir AUXILIAR_VDS.xlsx: {exc}")

    # Construir leitores indexados uma única vez (compartilhados entre todas as regiões)
    leitor_aux = LeitorAuxiliar(wsAux, "AUXILIAR")
    leitor_vds = LeitorAuxiliar(wsVD, "AUXILIAR_VDS")
    logger.info("AUXILIAR indexado: %d produtos, %d lojas",
                len(leitor_aux.produtos), len(leitor_aux.lojas))
    logger.info("AUXILIAR_VDS indexado: %d produtos, %d lojas",
                len(leitor_vds.produtos), len(leitor_vds.lojas))

    # Mapa de regiões
    regioes = {
        "praia": (resolve(cfg["planilhas_pai"]["praia"]), preencher_praia),
        "jau":   (resolve(cfg["planilhas_pai"]["jau"]),   preencher_jau),
        "bauru": (resolve(cfg["planilhas_pai"]["bauru"]), preencher_bauru),
    }

    alvo = [args.regiao] if args.regiao else list(regioes.keys())

    for regiao in alvo:
        caminho, preencher_fn = regioes[regiao]
        processar_planilha(regiao, caminho, leitor_aux, leitor_vds,
                           preencher_fn, args.dry_run, logger)

    # Avisos finais — produtos/lojas não encontrados nas auxiliares
    todas_faltantes = sorted(leitor_aux.faltantes | leitor_vds.faltantes)
    if todas_faltantes:
        print()
        print("=" * 60)
        print(f"  AVISOS — {len(todas_faltantes)} item(s) não encontrado(s) nas auxiliares")
        print("  (escrito 0 nessas células; revise as planilhas auxiliares)")
        print("=" * 60)
        for f in todas_faltantes:
            print(f"  - {f}")
            logger.warning(f)

    print("\n   Processo 1 concluído com sucesso.")
    print("=" * 60)
    logger.info("Processo 1 concluído.")


if __name__ == "__main__":
    main()
