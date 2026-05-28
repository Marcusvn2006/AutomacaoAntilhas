#!/usr/bin/env python3
"""
Processo 3 – Preenchimento de TOTAL ITENS DOADOS 2026
=======================================================
Para cada loja/CD/ER das planilhas pai (BAURU, PRAIA, JAÚ):
  1. Lê o valor USOU 1 de cada linha mapeada na aba da loja/CD/ER.
     USOU 1 = prev2 + Novo_Pedido - prev1
       prev2    = célula 2 colunas à esquerda do USOU (estoque semana anterior)
       prev1    = célula 1 coluna  à esquerda do USOU (estoque semana atual)
       Novo_Ped = OFFSET(USOU,0,6) se não-vazio, senão OFFSET(USOU,0,7)
  2. Encontra a coluna da semana atual (segunda-feira) na aba TOTAL ITENS DOADOS.
  3. Grava o resultado nas linhas correspondentes de cada bloco na aba TOTAL.

Uso:
    python processo3.py            # execução normal
    python processo3.py --dry-run  # simulação sem salvar
"""

import logging
import re
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Diretórios
# ---------------------------------------------------------------------------
BASE_DIR      = Path(__file__).resolve().parent
DIR_PLANILHAS = BASE_DIR / "04_PLANILHAS_PAI"
DIR_BACKUP    = BASE_DIR / "01_BACKUP"
DIR_LOGS      = BASE_DIR / "02_LOGS"
LOCK_FILE     = BASE_DIR / ".processo3.lock"

# ---------------------------------------------------------------------------
# Helpers de mapeamento
# ---------------------------------------------------------------------------

# Mapa fixo para lojas regulares (8 itens, offsets 2-9)
_MAPA_LOJA: List[Tuple[int, int]] = [
    (8,  2),   # SACOLA PP VERDE
    (9,  3),   # SACOLA P VERDE
    (10, 4),   # SACOLA M VERDE
    (11, 5),   # SACOLA G VERDE
    (14, 6),   # PAPEL DE SEDA
    (15, 7),   # ETIQUETAS
    (16, 8),   # TAGS
    (17, 9),   # LÂMINA OLFATIVA
]

def _pos(rows: List[int]) -> List[Tuple[int, int]]:
    """Mapeamento posicional: i-ésima linha → offset 2+i (slot sequencial no bloco TOTAL)."""
    return [(r, 2 + i) for i, r in enumerate(rows)]


# ---------------------------------------------------------------------------
# Configuração das planilhas pai, lojas, CDs e ERs
#
# Estrutura de cada entrada em "lojas":
#   nome_aba    : str  — nome exato da aba na planilha pai
#   linha_total : int  — linha do NOME da loja no bloco da aba TOTAL
#                        (offset 0 = nome, 1 = data, 2+ = itens)
#   mapa        : List[(store_row, total_offset)]
#                        — pares (linha da aba da loja, offset no bloco TOTAL)
# ---------------------------------------------------------------------------
PLANILHAS: List[dict] = [
    # ======================================================================
    {
        "nome": "BAURU",
        "arquivo": "BAURU - Contagem Antilhas - 2026.xlsm",
        "aba_total": "TOTAL ITENS DOADOS 2026",
        "lojas": [
            # --- Lojas regulares (8 itens) ---
            {"nome_aba": "BSH 6700",   "linha_total": 2,   "mapa": _MAPA_LOJA},
            {"nome_aba": "BOUL 13868", "linha_total": 15,  "mapa": _MAPA_LOJA},
            {"nome_aba": "TT 13370",   "linha_total": 28,  "mapa": _MAPA_LOJA},
            {"nome_aba": "TDQ 20000",  "linha_total": 41,  "mapa": _MAPA_LOJA},
            {"nome_aba": "GET 23049",  "linha_total": 54,  "mapa": _MAPA_LOJA},
            {"nome_aba": "Q7 6727",    "linha_total": 67,  "mapa": _MAPA_LOJA},
            {"nome_aba": "Q2 12466",   "linha_total": 80,  "mapa": _MAPA_LOJA},
            {"nome_aba": "MD 12942",   "linha_total": 93,  "mapa": _MAPA_LOJA},
            # --- CD (14 itens, offsets 2-15) ---
            {"nome_aba": "CDB - 23280", "linha_total": 106, "mapa": _pos([
                8, 9, 10, 11, 39, 47, 48, 49, 55, 56, 57, 66, 67, 68,
            ])},
            # --- ER (27 itens, offsets 2-28) ---
            {"nome_aba": "ERB - 22851", "linha_total": 125, "mapa": _pos([
                8, 9, 10, 11, 12, 13, 16, 17, 18, 19,
                44, 45, 46, 47,
                55, 56, 57, 58, 59, 60, 61,
                67, 68, 69, 70, 71, 72,
            ])},
        ],
    },

    # ======================================================================
    {
        "nome": "PRAIA",
        "arquivo": "PRAIA - Contagem Antilhas - 2026.xlsm",
        "aba_total": "TOTAL DE ITENS DOADOS 2026",   # "DE" intencional
        "lojas": [
            # --- Lojas regulares ---
            {"nome_aba": " PL 12973",   "linha_total": 2,   "mapa": _MAPA_LOJA},  # espaço inicial intencional
            {"nome_aba": "BOQ 11734",   "linha_total": 15,  "mapa": _MAPA_LOJA},
            {"nome_aba": "TP 14462",    "linha_total": 28,  "mapa": _MAPA_LOJA},
            {"nome_aba": "PB 5418",     "linha_total": 41,  "mapa": _MAPA_LOJA},
            {"nome_aba": "MG 11733",    "linha_total": 54,  "mapa": _MAPA_LOJA},
            {"nome_aba": "ATC - 23012", "linha_total": 67,  "mapa": _MAPA_LOJA},
            # --- CD ---
            {"nome_aba": "CDP - 24790",   "linha_total": 80,  "mapa": _pos([
                8, 9, 10, 11, 45, 53, 54, 55, 61, 62, 63, 72, 73, 74,
            ])},
            # --- ERs (27 itens cada) ---
            {"nome_aba": "ER BOQ - 23614", "linha_total": 99,  "mapa": _pos([
                8, 9, 10, 11, 12, 13, 16, 17, 18, 19,
                49, 50, 51, 52,
                60, 61, 62, 63, 64, 65, 66,
                72, 73, 74, 75, 76, 77,
            ])},
            {"nome_aba": "ER PBE - 23343", "linha_total": 131, "mapa": _pos([
                8, 9, 10, 11, 12, 13, 16, 17, 18, 19,
                49, 50, 51, 52,
                60, 61, 62, 63, 64, 65, 66,
                72, 73, 74, 75, 76, 77,
            ])},
            {"nome_aba": "ER MG - 24119",  "linha_total": 163, "mapa": _pos([
                8, 9, 10, 11, 12, 13, 16, 17, 18, 19,
                44, 45, 46, 47,
                55, 56, 57, 58, 59, 60, 61,
                67, 68, 69, 70, 71, 72,
            ])},
        ],
    },

    # ======================================================================
    {
        "nome": "JAU",
        "arquivo": "JAÚ - Contagem Antilhas - 2026.xlsm",
        "aba_total": "TOTAL ITENS DOADOS 2026",
        "lojas": [
            # --- Lojas regulares ---
            {"nome_aba": "JC 3822",    "linha_total": 2,   "mapa": _MAPA_LOJA},
            {"nome_aba": "JD 14446",   "linha_total": 15,  "mapa": _MAPA_LOJA},
            {"nome_aba": "BB 12066",   "linha_total": 28,  "mapa": _MAPA_LOJA},
            {"nome_aba": "SM 23048",   "linha_total": 41,  "mapa": _MAPA_LOJA},
            {"nome_aba": "JSH 11722",  "linha_total": 54,  "mapa": _MAPA_LOJA},
            {"nome_aba": "CONF 14553", "linha_total": 67,  "mapa": _MAPA_LOJA},
            {"nome_aba": "DC 7529",    "linha_total": 80,  "mapa": _MAPA_LOJA},
            {"nome_aba": "IT 6942",    "linha_total": 93,  "mapa": _MAPA_LOJA},
            {"nome_aba": "BR 6954",    "linha_total": 106, "mapa": _MAPA_LOJA},
            # --- CD ---
            {"nome_aba": "CDJ 23091",  "linha_total": 119, "mapa": _pos([
                8, 9, 10, 11, 36, 44, 45, 46, 52, 53, 54, 63, 64, 65,
            ])},
            # --- ER JAÚ (27 itens) ---
            {"nome_aba": "ERJ 22838",  "linha_total": 138, "mapa": _pos([
                8, 9, 10, 11, 12, 13, 16, 17, 18, 19,
                49, 50, 51, 52,
                60, 61, 62, 63, 64, 65, 66,
                72, 73, 74, 75, 76, 77,
            ])},
            # --- ER SÃO MANUEL (24 itens — sem G VERDE e sem KRAFT M/G)
            #     Mapeamento semântico: offsets 5,6,7 ficam vazios no TOTAL.
            #     Cada linha vai para o slot do seu produto, não por posição.
            {"nome_aba": "ER SM 24137", "linha_total": 170, "mapa": [
                (8,  2),  # SACOLA PP VERDE
                (9,  3),  # SACOLA P VERDE
                (10, 4),  # SACOLA M VERDE
                # offset 5 = G VERDE    → não existe no SM
                # offset 6 = KRAFT M    → não existe no SM
                # offset 7 = KRAFT G    → não existe no SM
                (16, 8),  (17, 9),  (18, 10), (19, 11),  # PAPEL/ETIQ/TAG/LÂM
                (44, 12), (45, 13), (46, 14), (47, 15),  # OUI (PAPEL/ETIQ/LÂM/SACO)
                (55, 16), (56, 17), (57, 18), (58, 19),  # QDB (SACOLA P/M/G + PAPEL)
                (59, 20), (60, 21), (61, 22),             # QDB (ETIQ/TAG/LÂM)
                (67, 23), (68, 24), (69, 25),             # EUDORA (SACOLA P/M/G)
                (70, 26), (71, 27), (72, 28),             # EUDORA (PAPEL/ADESIVOS/LÂM)
            ]},
        ],
    },
]


# ---------------------------------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------------------------------


def _is_blank(val) -> bool:
    """Verifica se a célula está de fato vazia (None ou string vazia)."""
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False

def _n_excel(val) -> float:
    """Replica N() do Excel: tenta converter para número, senão vira 0."""
    if val is None or isinstance(val, bool):
        return 0.0
    try:
        # Força a conversão para float para garantir que strings como "0" sejam lidas como número
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def _encontrar_usou_col(ws, header_row: int = 6) -> Optional[int]:
    """Localiza a coluna cujo cabeçalho é exatamente 'USOU' na linha indicada."""
    row_vals = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    for col_idx, val in enumerate(row_vals, start=1):
        if val is not None and str(val).strip().upper() == "USOU":
            return col_idx
    return None


def _encontrar_coluna_semana(ws, data_row: int = 3) -> Optional[int]:
    """Encontra a coluna da segunda-feira da semana atual na linha de datas."""
    hoje   = date.today()
    segunda = hoje - timedelta(days=hoje.weekday())
    row_vals = list(ws.iter_rows(min_row=data_row, max_row=data_row, values_only=True))[0]
    for col_idx, val in enumerate(row_vals, start=1):
        if isinstance(val, datetime) and val.date() == segunda:
            return col_idx
        if isinstance(val, date) and not isinstance(val, datetime) and val == segunda:
            return col_idx
    return None


def _parse_usou_formula(formula_str: str) -> Tuple[int, int]:
    """
    Lê a fórmula USOU armazenada no XML do Excel (sempre em inglês) e extrai
    os offsets de NP_A e NP_B a partir dos dois N(OFFSET(...,0,X)) presentes.

    Exemplos armazenados pelo Excel:
      =BZ8+IF(NOT(ISBLANK(OFFSET(CB8,0,6))),N(OFFSET(CB8,0,6)),N(OFFSET(CB8,0,7)))-CA8
      =AB8+IF(NOT(ISBLANK(OFFSET(AD8,0,7))),N(OFFSET(AD8,0,7)),N(OFFSET(AD8,0,8))-AC8)

    Fallback seguro: offsets 6 e 7.
    """
    if not formula_str or not isinstance(formula_str, str) or not formula_str.startswith("="):
        return 6, 7  # fallback

    f = formula_str.upper().strip()

    # Pega os offsets apenas dos N(OFFSET(...,0,X)) — ignora o OFFSET dentro do ISBLANK
    np_offsets = re.findall(r"N\(OFFSET\([A-Z]+\d+,0,(\d+)\)\)", f)
    if len(np_offsets) >= 2:
        return int(np_offsets[0]), int(np_offsets[1])

    # Fallback: pega qualquer OFFSET com offsets únicos em ordem
    all_off = list(dict.fromkeys(re.findall(r"OFFSET\([A-Z]+\d+,0,(\d+)\)", f)))
    np_a_off = int(all_off[0]) if len(all_off) > 0 else 6
    np_b_off = int(all_off[1]) if len(all_off) > 1 else np_a_off + 1
    return np_a_off, np_b_off



def _compute_usou(ws, row_idx: int, usou_col: int,
                  np_a_off: int, np_b_off: int) -> int:
    """
    Calcula o valor USOU validando corretamente o zero:
      = prev2 + IF(NOT(ISBLANK(NP_A)), N(NP_A), N(NP_B)) - prev1
    """
    if usou_col < 3:
        return 0
    
    prev2_val = ws.cell(row=row_idx, column=usou_col - 2).value
    prev1_val = ws.cell(row=row_idx, column=usou_col - 1).value
    np_a_raw = ws.cell(row=row_idx, column=usou_col + np_a_off).value
    np_b_raw = ws.cell(row=row_idx, column=usou_col + np_b_off).value

    prev2 = _n_excel(prev2_val)
    prev1 = _n_excel(prev1_val)

    if not _is_blank(np_a_raw):
        # NP_A presente (inclusive se for 0): mantém o valor dele e subtrai o estoque atual
        resultado = prev2 + _n_excel(np_a_raw) - prev1
    else:
        # NP_A totalmente vazio: pula para o NP_B e subtrai o estoque atual
        resultado = prev2 + _n_excel(np_b_raw) - prev1

    return int(round(resultado))
# ---------------------------------------------------------------------------
# Processamento de uma planilha pai
# ---------------------------------------------------------------------------

def processar_planilha(cfg: dict, dry_run: bool, logger: logging.Logger) -> Dict[str, int]:
    """
    Processa uma planilha pai completa.
    Retorna dict {nome_aba: nº_de_células_escritas}.
    """
    contagem: Dict[str, int] = {}
    filepath = DIR_PLANILHAS / cfg["arquivo"]

    if not filepath.exists():
        logger.error("[%s] Arquivo nao encontrado: %s", cfg["nome"], filepath)
        return contagem

    logger.info("[%s] Abrindo: %s", cfg["nome"], cfg["arquivo"])
    # wb        → fórmulas em texto + destino de escrita no TOTAL
    # wb_data   → valores calculados (data_only=True) para leitura das abas de loja
    wb      = openpyxl.load_workbook(str(filepath), keep_vba=True)
    wb_data = openpyxl.load_workbook(str(filepath), keep_vba=True, data_only=True)

    aba_total = cfg["aba_total"]
    if aba_total not in wb.sheetnames:
        logger.error("[%s] Aba '%s' nao encontrada. Abas: %s",
                     cfg["nome"], aba_total, wb.sheetnames)
        return contagem

    ws_total = wb[aba_total]
    week_col = _encontrar_coluna_semana(ws_total, data_row=3)
    if week_col is None:
        hoje    = date.today()
        segunda = hoje - timedelta(days=hoje.weekday())
        logger.error("[%s] Semana %s nao encontrada em '%s'.",
                     cfg["nome"], segunda.strftime("%d/%m/%Y"), aba_total)
        return contagem

    hoje    = date.today()
    segunda = hoje - timedelta(days=hoje.weekday())
    logger.info("[%s] Semana %s -> coluna %s",
                cfg["nome"], segunda.strftime("%d/%m/%Y"), get_column_letter(week_col))

    for loja in cfg["lojas"]:
        nome_aba    = loja["nome_aba"]
        linha_total = loja["linha_total"]
        mapa        = loja["mapa"]

        if nome_aba not in wb.sheetnames:
            logger.warning("[%s] Aba '%s' nao encontrada -- pulada.", cfg["nome"], nome_aba)
            continue

        # ws_loja      → fórmulas (para parsear a estrutura do USOU)
        # ws_loja_data → valores calculados (para ler prev2/prev1/NP_A/NP_B)
        ws_loja      = wb[nome_aba]
        ws_loja_data = wb_data[nome_aba]

        usou_col = _encontrar_usou_col(ws_loja, header_row=6)
        if usou_col is None:
            logger.warning("[%s] '%s': coluna USOU nao encontrada na row 6 -- pulada.",
                           cfg["nome"], nome_aba)
            continue

        # Parsear a fórmula USOU (lida do wb de fórmulas, uma vez por aba)
        formula_raw = ws_loja.cell(row=mapa[0][0], column=usou_col).value
        np_a_off, np_b_off = _parse_usou_formula(
            str(formula_raw) if formula_raw is not None else ""
        )
        logger.info("[%s] %-16s: USOU=%-3s  linha_total=%d  itens=%d  NP_A=+%d NP_B=+%d",
                    cfg["nome"], nome_aba.strip(), get_column_letter(usou_col),
                    linha_total, len(mapa), np_a_off, np_b_off)

        for store_row, total_offset in mapa:
            total_row = linha_total + total_offset
            # Calcula com ws_loja_data (valores reais, não strings de fórmula)
            usou_val  = _compute_usou(ws_loja_data, store_row, usou_col,
                                      np_a_off, np_b_off)
            logger.debug("    store_row=%-2d  total_row=%-3d  col=%-3s  val=%d",
                         store_row, total_row, get_column_letter(week_col), usou_val)
            if not dry_run:
                ws_total.cell(row=total_row, column=week_col).value = usou_val

        contagem[nome_aba] = len(mapa)

    if not dry_run:
        wb.save(str(filepath))
        logger.info("[%s] Arquivo salvo.", cfg["nome"])
    else:
        logger.info("[%s] [DRY-RUN] Nenhuma alteracao salva.", cfg["nome"])

    return contagem


# ---------------------------------------------------------------------------
# Log / Backup / Lock
# ---------------------------------------------------------------------------

def _configurar_log(dry_run: bool) -> logging.Logger:
    DIR_LOGS.mkdir(parents=True, exist_ok=True)
    hoje_str = date.today().strftime("%Y-%m-%d")
    log_path = DIR_LOGS / f"processo3_{hoje_str}.log"

    logger = logging.getLogger("processo3")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")

    fh = logging.FileHandler(str(log_path), encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


def _fazer_backup(logger: logging.Logger) -> None:
    DIR_BACKUP.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y-%m-%d_%Hh%M_p3")
    dest = DIR_BACKUP / ts
    dest.mkdir(exist_ok=True)
    copiados = 0
    for cfg in PLANILHAS:
        src = DIR_PLANILHAS / cfg["arquivo"]
        if src.exists():
            shutil.copy2(str(src), str(dest / cfg["arquivo"]))
            copiados += 1
    logger.info("Backup criado em '%s' (%d arquivo(s)).", ts, copiados)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(dry_run: bool = False) -> None:
    logger = _configurar_log(dry_run)

    if LOCK_FILE.exists():
        logger.error("ERRO: .processo3.lock ja existe -- outra execucao em andamento.")
        sys.exit(1)

    try:
        LOCK_FILE.write_text("running", encoding="utf-8")

        logger.info("=" * 60)
        logger.info("PROCESSO 3 -- TOTAL ITENS DOADOS 2026")
        if dry_run:
            logger.info("[DRY-RUN] Nenhuma alteracao sera salva.")
        logger.info("=" * 60)

        if not dry_run:
            _fazer_backup(logger)

        total_geral = 0
        for cfg in PLANILHAS:
            contagem       = processar_planilha(cfg, dry_run, logger)
            total_planilha = sum(contagem.values())
            total_geral   += total_planilha
            logger.info("[%s] %d celula(s) em %d loja(s)/CD/ER.",
                        cfg["nome"], total_planilha, len(contagem))

        logger.info("=" * 60)
        logger.info("CONCLUIDO: %d celula(s) no total.", total_geral)
        logger.info("=" * 60)

    finally:
        if LOCK_FILE.exists():
            LOCK_FILE.unlink()


if __name__ == "__main__":
    _dry = "--dry-run" in sys.argv or "-n" in sys.argv
    main(dry_run=_dry)
