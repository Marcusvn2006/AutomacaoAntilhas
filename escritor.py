"""
escritor.py — Escrita dos valores na Planilha Pai (.xlsm).

Responsabilidades:
  - Abrir a Planilha Pai com keep_vba=True (preserva macros)
  - Localizar a coluna correta pela data na linha 6
  - Localizar a linha correta pelo nome do produto na coluna A
  - Escrever o valor na interseção
  - Salvar e fechar preservando as macros

Regras:
  - Comparações de texto sempre com .strip().upper() nos dois lados
  - Coluna não encontrada → ErroEscrita (loja vai para quarentena)
  - Produto não encontrado na aba → aviso no log, continua os demais
"""

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Linha onde ficam as datas das semanas na Planilha Pai
LINHA_DATAS = 6

# Coluna onde ficam os nomes dos produtos na Planilha Pai
COL_PRODUTO = 1  # coluna A


class ErroEscrita(Exception):
    """Erro ao tentar escrever na Planilha Pai — impede o processamento da loja."""


def escrever_na_planilha_pai(
    caminho_pai: Path,
    nome_aba: str,
    data_contagem: date,
    dados: dict[str, Any],
    mapa_produtos: list[dict],
) -> None:
    """
    Escreve os dados de uma loja na aba correspondente da Planilha Pai.

    Args:
        caminho_pai:    Caminho completo para o arquivo .xlsm.
        nome_aba:       Nome da aba da loja (ex: 'JC 3822').
        data_contagem:  Data da contagem semanal.
        dados:          {nome_loja_upper: valor} retornado pelo processador.
        mapa_produtos:  Lista de produtos do config.yaml (para tradução de nomes).

    Raises:
        ErroEscrita: se o arquivo não puder ser aberto, a aba não existir
                     ou a coluna da data não for encontrada.
    """
    try:
        wb = openpyxl.load_workbook(caminho_pai, keep_vba=True)
    except Exception as exc:
        raise ErroEscrita(
            f"Não foi possível abrir a Planilha Pai '{caminho_pai.name}': {exc}"
        ) from exc

    if nome_aba not in wb.sheetnames:
        raise ErroEscrita(
            f"Aba '{nome_aba}' não encontrada em '{caminho_pai.name}'. "
            f"Abas disponíveis: {wb.sheetnames}"
        )

    ws = wb[nome_aba]

    col_semana = _encontrar_coluna_data(ws, data_contagem, caminho_pai.name, nome_aba)

    produtos_escritos = 0
    produtos_nao_encontrados = []

    for produto_cfg in mapa_produtos:
        nome_loja  = produto_cfg["nome_loja"].strip().upper()
        nome_pai   = produto_cfg["nome_pai"].strip().upper()

        valor = dados.get(nome_loja)
        if valor is None:
            logger.warning(
                "[%s] Valor None para '%s' — célula não será preenchida.", nome_aba, nome_loja
            )
            continue

        linha_produto = _encontrar_linha_produto(ws, nome_pai)
        if linha_produto is None:
            produtos_nao_encontrados.append(nome_pai)
            continue

        ws.cell(row=linha_produto, column=col_semana).value = valor
        logger.debug(
            "[%s] Escrito: '%s' → col=%s linha=%d valor=%s",
            nome_aba, nome_pai, get_column_letter(col_semana), linha_produto, valor,
        )
        produtos_escritos += 1

    if produtos_nao_encontrados:
        logger.warning(
            "[%s] Produtos não encontrados na aba (não escritos): %s",
            nome_aba, produtos_nao_encontrados,
        )

    try:
        wb.save(caminho_pai)
    except Exception as exc:
        raise ErroEscrita(
            f"Erro ao salvar '{caminho_pai.name}' após escrita: {exc}"
        ) from exc

    logger.info(
        "[%s] %d produto(s) escritos para %s.", nome_aba, produtos_escritos, data_contagem
    )


# ── Auxiliares ──────────────────────────────────────────────────────────────

def _encontrar_coluna_data(
    ws, data_contagem: date, nome_arquivo: str, nome_aba: str
) -> int:
    """
    Varre a linha 6 procurando a coluna cuja célula corresponda à data_contagem.
    Aceita células do tipo date, datetime ou string ISO.

    Raises:
        ErroEscrita: se a data não for encontrada em nenhuma coluna.
    """
    for col in range(1, ws.max_column + 1):
        celula = ws.cell(row=LINHA_DATAS, column=col)
        data_celula = _normalizar_data(celula.value)
        if data_celula == data_contagem:
            logger.debug(
                "[%s] Coluna da data %s encontrada: %s (col %d)",
                nome_aba, data_contagem, get_column_letter(col), col,
            )
            return col

    raise ErroEscrita(
        f"[{nome_aba}] Data {data_contagem} não encontrada na linha {LINHA_DATAS} "
        f"de '{nome_arquivo}'. Verifique se a semana já foi adicionada à planilha."
    )


def _encontrar_linha_produto(ws, nome_pai_upper: str) -> int | None:
    """
    Varre a coluna A procurando o produto pelo nome (comparação case-insensitive sem espaços).
    Começa na linha 7 (após cabeçalhos).
    Retorna o número da linha ou None se não encontrar.
    """
    for row in range(7, ws.max_row + 1):
        celula = ws.cell(row=row, column=COL_PRODUTO)
        if celula.value is None:
            continue
        nome_celula = str(celula.value).strip().upper()
        if nome_celula == nome_pai_upper:
            return row
    return None


def _normalizar_data(valor: Any) -> date | None:
    """
    Converte o valor de uma célula para objeto date.
    Aceita: date, datetime, string 'AAAA-MM-DD' ou 'DD/MM/AAAA'.
    Retorna None se não conseguir converter.
    """
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(texto, fmt).date()
            except ValueError:
                continue
    return None
