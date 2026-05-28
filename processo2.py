"""
Processo 2 — Integração das planilhas enviadas pelas lojas nas Planilhas Pai.

Lê arquivos .xlsm/.xlsx em 00_ENTRADA/AUXILIAR_EMAIL/<REGIAO>/PENDENTES/,
integra os valores na planilha pai correspondente (mesma coluna USOU criada
pelo Processo 1, em linhas diferentes), e move os arquivos processados
para PROCESSADOS/<data>/.

Uso:
    python processo2.py                    # execução normal (todas as regiões)
    python processo2.py --dry-run          # simula sem mover/salvar
    python processo2.py --regiao bauru     # processa só Bauru
"""
from __future__ import annotations

import argparse
import datetime as dt
import gc
import json
import logging
import math
import re
import shutil
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import yaml

# ─────────────────────────────────────────────────────────────────────────────
# PATCH: openpyxl 3.1.5 ainda não suporta o atributo `extLst` que Excel 365
# grava dentro de <patternFill>. Como só lemos dados (não formatação),
# fazemos PatternFill IGNORAR esse kwarg para não crashar com arquivos novos.
# Bug: https://foss.heptapod.net/openpyxl/openpyxl/-/issues (extLst PatternFill)
# ─────────────────────────────────────────────────────────────────────────────
from openpyxl.styles.fills import PatternFill as _PatternFill
_orig_PatternFill_init = _PatternFill.__init__

def _safe_PatternFill_init(self, *args, **kwargs):
    kwargs.pop("extLst", None)  # descarta atributo não suportado
    _orig_PatternFill_init(self, *args, **kwargs)

_PatternFill.__init__ = _safe_PatternFill_init


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

LOCK_FILE_NAME = ".processo2.lock"
EXTENSOES_ACEITAS = (".xlsm", ".xlsx")
# Padrão de "arquivo duplicado" comum em downloads: "Nome (1).xlsm", "Nome (2).xlsx"
_RE_DUPLICATA = re.compile(r"\s*\(\d+\)\s*$")


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _normalizar_sigla(nome_arquivo: str) -> str:
    """'BSH.xlsm' → 'BSH'  |  'bsh.xlsx' → 'BSH'  |  'BSH (1).xlsm' → 'BSH (1)'"""
    p = Path(nome_arquivo)
    return p.stem.strip().upper()


def _eh_duplicata(nome_arquivo: str) -> bool:
    """Detecta padrão 'Nome (1).xlsm', 'Nome (2).xlsx', etc."""
    return bool(_RE_DUPLICATA.search(Path(nome_arquivo).stem))


def _eh_numerico(valor) -> bool:
    """True se for int, float ou bool (mas não strings). None retorna False."""
    if valor is None:
        return False
    if isinstance(valor, bool):
        return False  # bool é subclasse de int, mas não queremos True/False
    return isinstance(valor, (int, float))


_NUMERO_RE = re.compile(r"-?\d+(?:[.,]\d+)?")


def _extrair_numero(valor):
    """
    Extrai um número de qualquer valor de célula.
      - None             → None
      - int / float      → float(valor)
      - bool             → None  (evita True=1, False=0)
      - str com número   → primeiro número encontrado (ex.: '7 PCT' → 7.0)
      - str sem número   → None
    Aceita vírgula ou ponto como decimal.
    """
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str):
        s = valor.strip()
        if not s:
            return None
        m = _NUMERO_RE.search(s)
        if m:
            return float(m.group().replace(",", "."))
    return None


def _cel_tem_data(cell) -> bool:
    """Retorna True se a célula contém uma data (objeto date/datetime)."""
    val = cell.value
    if val is None:
        return False
    return isinstance(val, (dt.date, dt.datetime))


def _mesma_semana(d1: dt.date, d2: dt.date) -> bool:
    """True se as duas datas caem na mesma semana ISO (segunda a domingo)."""
    return d1.isocalendar()[:2] == d2.isocalendar()[:2]


def _aplicar_regra_pct(valor: float, pct: Optional[float]) -> float:
    """
    Converte unidades em caixas usando o PCT (unidades por caixa).
    Aplicado apenas em células com `pct_de` configurado no YAML.

    Regra (escopo: tabela DIVERSOS BOTICÁRIO 2025 — PAPEL DE SEDA, ETIQUETAS,
    TAG e LÂMINA OLFATIVA):
        PCT inválido (None, 0 ou negativo)  →  retorna V (pula conversão)
        V <= 30                             →  retorna V    (já está em caixas)
        30 < V < PCT                        →  retorna 0    (fração de caixa)
        V >= PCT                            →  retorna floor(V / PCT)
    """
    if pct is None or pct <= 0:
        return valor
    if valor <= 30:
        return valor
    if valor < pct:
        return 0.0
    return float(math.floor(valor / pct))


def _aplicar_regra_pct_mult(valor: float, pct: Optional[float]) -> float:
    """
    Converte caixas em unidades usando o PCT (unidades por caixa).
    Inverso de _aplicar_regra_pct — usado quando a loja informa em caixas
    mas o correto é em unidades (chave `pct_mult` no YAML).

    Regra:
        PCT inválido (None, 0 ou negativo) → retorna V (sem conversão)
        V ≥ 50                             → retorna V (já está em unidades)
        PCT ≤ V < 50                       → retorna V (em unidades, pode ter decimal)
        V < PCT (e V < 50)                 → retorna floor(V × PCT) (caixas → unidades)
    """
    if pct is None or pct <= 0:
        return valor
    if valor >= 50:
        return valor
    if valor >= pct:
        return valor  # entre PCT e 50: já em unidades (pode ter decimal)
    return float(math.floor(valor * pct))


# ─────────────────────────────────────────────────────────────────────────────
# DETECÇÃO DE COLUNAS USOU NA PLANILHA PAI
# ─────────────────────────────────────────────────────────────────────────────

def _encontrar_cols_usou_old(ws) -> List[int]:
    """Retorna lista 1-based de colunas que contêm o texto 'USOU' (linhas 1-30).
    Essas são as colunas USOU *originais* (não as novas criadas pelo Processo 1).
    """
    resultado = []
    for col in range(1, ws.max_column + 1):
        for row in range(1, 31):
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).strip().upper() == "USOU":
                resultado.append(col)
                break
    return resultado


def encontrar_cols_usou_novas(ws) -> Dict[int, int]:
    """
    Retorna {índice_usou: coluna_nova} (1-based, esquerda→direita).
    A nova coluna do Processo 1 fica imediatamente à esquerda de cada USOU original.
    Ex: {1: 79, 2: 180} para BSH 6700.
    """
    cols_old = sorted(_encontrar_cols_usou_old(ws))
    return {i: col - 1 for i, col in enumerate(cols_old, start=1)}


def detectar_data_processo1(ws, max_row_scan: int = 100) -> Optional[dt.date]:
    """
    Detecta a data da última execução do Processo 1.
    Olha a coluna nova de USOU 1 (esquerda) e procura a maior data escrita lá.
    Para performance, escaneia no máximo `max_row_scan` linhas.
    Retorna None se não houver USOU ou nenhuma data.
    """
    novas = encontrar_cols_usou_novas(ws)
    if 1 not in novas:
        return None
    col_nova_usou1 = novas[1]
    datas: List[dt.date] = []
    limite = min(ws.max_row, max_row_scan)
    for row in range(1, limite + 1):
        val = ws.cell(row=row, column=col_nova_usou1).value
        if isinstance(val, dt.datetime):
            datas.append(val.date())
        elif isinstance(val, dt.date):
            datas.append(val)
    return max(datas) if datas else None


# ─────────────────────────────────────────────────────────────────────────────
# BACKUP
# ─────────────────────────────────────────────────────────────────────────────

def fazer_backup(planilhas_pai: Dict[str, Path], pasta_backup: Path,
                 dry_run: bool, logger: logging.Logger) -> Path:
    """Copia as planilhas pai para 01_BACKUP/AAAA-MM-DD_HHhMM/."""
    agora = dt.datetime.today()
    nome_pasta = agora.strftime("%Y-%m-%d_%Hh%M") + "_p2"
    destino = pasta_backup / nome_pasta

    if dry_run:
        logger.info("[DRY-RUN] Backup seria criado em: %s", destino)
        print(f"  [DRY-RUN] Backup seria criado em: {destino}")
        return destino

    try:
        destino.mkdir(parents=True, exist_ok=True)
        for regiao, caminho in planilhas_pai.items():
            if not caminho.exists():
                raise FileNotFoundError(f"Planilha pai não encontrada: {caminho}")
            shutil.copy2(caminho, destino / caminho.name)
            logger.debug("  Backup: %s", caminho.name)
        print(f"  Backup criado em: {destino}")
        logger.info("Backup criado em: %s", destino)
        return destino
    except Exception as exc:
        logger.critical("Falha no backup: %s", exc)
        sys.exit(f"ERRO CRÍTICO: Backup falhou — {exc}")


# ─────────────────────────────────────────────────────────────────────────────
# MOVIMENTAÇÃO DE ARQUIVOS
# ─────────────────────────────────────────────────────────────────────────────

def mover_para_processados(arquivo: Path, pasta_regiao: Path,
                           data_str: str, dry_run: bool,
                           logger: logging.Logger) -> Path:
    """Move o arquivo para PROCESSADOS/<data>/. Retorna o caminho final."""
    destino_dir = pasta_regiao / "PROCESSADOS" / data_str
    destino = destino_dir / arquivo.name
    if dry_run:
        logger.debug("[DRY-RUN] Moveria %s → %s", arquivo, destino)
        return destino
    destino_dir.mkdir(parents=True, exist_ok=True)
    for _tentativa in range(5):
        try:
            shutil.move(str(arquivo), str(destino))
            logger.info("Movido: %s → %s", arquivo.name, destino)
            return destino
        except PermissionError:
            if _tentativa < 4:
                time.sleep(1.0)
    # Não conseguiu mover — avisa mas não crasha
    logger.warning(
        "AVISO: não foi possível mover %s para PROCESSADOS (arquivo bloqueado). "
        "Arquivo permanece em PENDENTES.",
        arquivo.name,
    )
    return arquivo


def mover_para_erros(arquivo: Path, pasta_regiao: Path, data_str: str,
                     sigla: str, motivo: str, detalhe: str,
                     dry_run: bool, logger: logging.Logger) -> Tuple[Path, Path]:
    """Move o arquivo para ERROS/<data>/ e cria <SIGLA>_ERRO.txt ao lado."""
    destino_dir = pasta_regiao / "ERROS" / data_str
    destino_arq = destino_dir / arquivo.name
    destino_txt = destino_dir / f"{sigla}_ERRO.txt"

    conteudo_txt = (
        f"LOJA: {sigla}\n"
        f"DATA: {dt.datetime.today().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"ARQUIVO: {arquivo.name}\n"
        f"MOTIVO: {motivo}\n"
        f"DETALHE: {detalhe}\n"
        f"\n"
        f"→ Corrija o arquivo e coloque de volta em PENDENTES/\n"
    )

    if dry_run:
        logger.debug("[DRY-RUN] Moveria %s → %s (+ %s)", arquivo, destino_arq, destino_txt.name)
        return destino_arq, destino_txt

    destino_dir.mkdir(parents=True, exist_ok=True)
    # Retry: antivírus/indexador pode manter o arquivo aberto por alguns segundos.
    # 5 tentativas × 1 s = até 4 s de espera.
    movido = False
    for _tentativa in range(5):
        try:
            shutil.move(str(arquivo), str(destino_arq))
            movido = True
            break
        except PermissionError:
            if _tentativa < 4:
                time.sleep(1.0)

    if not movido:
        # Arquivo bloqueado mesmo após retries (AV, indexador, outro processo).
        # Deixa em PENDENTES e grava o txt de erro ao lado — NÃO crashar o script.
        logger.warning(
            "AVISO: não foi possível mover %s para ERROS (arquivo bloqueado). "
            "Deixando em PENDENTES. Motivo do erro original: %s",
            arquivo.name, motivo,
        )
        txt_pendentes = arquivo.parent / f"{sigla}_ERRO.txt"
        txt_pendentes.write_text(conteudo_txt, encoding="utf-8")
        return arquivo, txt_pendentes

    destino_txt.write_text(conteudo_txt, encoding="utf-8")
    logger.warning("ERRO: %s → %s (%s)", arquivo.name, destino_arq, motivo)
    return destino_arq, destino_txt


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DE UM ARQUIVO DE LOJA
# ─────────────────────────────────────────────────────────────────────────────

def ler_arquivo_loja(
    arquivo: Path,
    aba_loja_esperada: str,
    destinos: Dict[str, dict],
    logger: logging.Logger,
) -> Tuple[Optional[Dict[str, float]], Optional[str], Optional[str]]:
    """
    Lê um arquivo .xlsm/.xlsx da loja e retorna o buffer de valores.

    Returns:
        (buffer, motivo_erro, detalhe_erro)
        - buffer: dict {celula_origem: valor} com células a serem escritas
                  (apenas células com valor numérico)
        - motivo_erro: string curta descrevendo o erro (None se OK)
        - detalhe_erro: detalhes adicionais
    """
    try:
        # read_only=False: openpyxl carrega tudo na memória e fecha o handle
        # do disco imediatamente após o load — sem isso o ZipFile fica aberto
        # durante toda a leitura e bloqueia o shutil.move subsequente.
        wb = openpyxl.load_workbook(str(arquivo), data_only=True, keep_vba=False)
    except PermissionError:
        return None, "ARQUIVO_ABERTO", "Arquivo está aberto no Excel"
    except Exception as exc:
        return None, "ARQUIVO_CORROMPIDO", f"Falha ao abrir: {exc}"

    try:
        if aba_loja_esperada not in wb.sheetnames:
            return None, "ABA_AUSENTE", (
                f"Aba '{aba_loja_esperada}' não encontrada. "
                f"Abas no arquivo: {wb.sheetnames}"
            )

        ws = wb[aba_loja_esperada]
        buffer: Dict[str, float] = {}

        for celula_origem, destino_cfg in destinos.items():
            try:
                col_letra = ''.join(c for c in celula_origem if c.isalpha())
                linha = int(''.join(c for c in celula_origem if c.isdigit()))
                col_idx = column_index_from_string(col_letra)
            except Exception:
                return None, "CELULA_INVALIDA", f"Célula '{celula_origem}' tem formato inválido"

            ignorar_zerado = (
                isinstance(destino_cfg, dict) and destino_cfg.get("ignorar_zerado")
            )

            # ── pct_mult: fórmula (duas_antes × PCT) + uma_antes ─────────
            # Lê as duas colunas à esquerda da célula de destino (ex: para M6
            # lê K6 e L6), multiplica K6 pelo PCT e soma L6.
            # Não usa o valor da célula principal (ex: M6) — a fórmula é
            # calculada inteiramente aqui.
            pct_mult = destino_cfg.get("pct_mult") if isinstance(destino_cfg, dict) else None
            if pct_mult:
                try:
                    col_pct_letra = ''.join(c for c in pct_mult if c.isalpha())
                    linha_pct = int(''.join(c for c in pct_mult if c.isdigit()))
                    pct_raw = ws.cell(
                        row=linha_pct,
                        column=column_index_from_string(col_pct_letra),
                    ).value
                    pct_val_fm = _extrair_numero(pct_raw) or 0.0

                    duas_raw = ws.cell(row=linha, column=col_idx - 2).value
                    uma_raw  = ws.cell(row=linha, column=col_idx - 1).value

                    # Ambas as células de entrada vazias → sem dado, pula
                    if duas_raw is None and uma_raw is None:
                        continue

                    duas_val = _extrair_numero(duas_raw) or 0.0
                    uma_val  = _extrair_numero(uma_raw)  or 0.0

                    valor = _aplicar_regra_pct_mult(duas_val, pct_val_fm) + uma_val
                except Exception as exc:
                    logger.warning(
                        "Erro na fórmula pct_mult em %s: %s", celula_origem, exc
                    )
                    continue

                if ignorar_zerado and valor == 0.0:
                    logger.debug(
                        "Ignorando %s (ignorar_zerado: resultado=0)", celula_origem
                    )
                    continue

                logger.info(
                    "Fórmula pct_mult em %s: pct_mult(%g, %g) + %g = %g",
                    celula_origem, duas_val, pct_val_fm, uma_val, valor,
                )
                buffer[celula_origem] = valor
                continue  # célula processada — pula fluxo normal abaixo

            # ── Fluxo normal (sem pct_mult) ───────────────────────────────
            val_raw = ws.cell(row=linha, column=col_idx).value
            # Extrai número mesmo se a loja digitou texto junto (ex: '7 PCT' → 7.0)
            val = _extrair_numero(val_raw)

            # Regra "ignorar se zerado": valor None ou 0 → pula a célula
            if ignorar_zerado:
                if val is None or val == 0.0:
                    logger.debug(
                        "Ignorando %s (regra ignorar_zerado: V=%r)",
                        celula_origem, val_raw,
                    )
                    continue

            if val_raw is None:
                # Célula vazia: pula (não escreve na planilha pai)
                continue
            if val is None:
                # Tinha algo na célula mas sem nenhum número extraível
                return None, "VALOR_NAO_NUMERICO", (
                    f"Célula '{celula_origem}' contém valor não numérico: {val_raw!r}"
                )

            # Avisa quando extraiu número de uma string (ex: '7 PCT' → 7)
            if isinstance(val_raw, str):
                logger.info(
                    "Célula %s: extraído número %g de '%s'",
                    celula_origem, val, val_raw,
                )

            valor = val

            # ── Regra PCT: conversão unidades → caixas ────────────────────
            pct_de = destino_cfg.get("pct_de") if isinstance(destino_cfg, dict) else None
            if pct_de:
                pct_val: Optional[float] = None
                try:
                    col_pct_letra = ''.join(c for c in pct_de if c.isalpha())
                    linha_pct = int(''.join(c for c in pct_de if c.isdigit()))
                    col_pct_idx = column_index_from_string(col_pct_letra)
                    pct_raw = ws.cell(row=linha_pct, column=col_pct_idx).value
                    # Extrai número mesmo se for string tipo '12 PCT'
                    pct_val = _extrair_numero(pct_raw)
                except Exception:
                    pct_val = None

                valor_antes = valor
                valor = _aplicar_regra_pct(valor, pct_val)
                if valor != valor_antes:
                    logger.info(
                        "Conversão PCT em %s: V=%g, PCT=%s → %g caixas",
                        celula_origem, valor_antes,
                        f"{pct_val:g}" if pct_val is not None else "n/a",
                        valor,
                    )
                elif pct_val is None:
                    logger.info(
                        "PCT ausente/inválido em %s (cel %s): mantendo V=%g",
                        celula_origem, pct_de, valor,
                    )

            buffer[celula_origem] = valor

        return buffer, None, None
    finally:
        wb.close()
        del wb       # remove referência explícita
        gc.collect() # força GC: libera ExcelReader.archive (ZipFile) antes do move


# ─────────────────────────────────────────────────────────────────────────────
# PROCESSAMENTO DE UMA REGIÃO
# ─────────────────────────────────────────────────────────────────────────────

def processar_regiao(
    regiao: str,
    pasta_regiao: Path,
    caminho_planilha_pai: Path,
    cfg: dict,
    dry_run: bool,
    data_str: str,
    logger: logging.Logger,
) -> dict:
    """Processa todos os arquivos em PENDENTES/ de uma região.
    Retorna um dicionário resumo para o manifesto.
    """
    label = regiao.upper()
    resumo: dict = {
        "processadas": [],
        "com_erro": [],
        "pendentes": [],
    }

    pasta_pendentes = pasta_regiao / "PENDENTES"
    if not pasta_pendentes.exists():
        logger.warning("[%s] PENDENTES/ não existe: %s", label, pasta_pendentes)
        return resumo

    # ── 1. Lista arquivos válidos em PENDENTES/ ────────────────────────────
    arquivos_brutos = [
        f for f in pasta_pendentes.iterdir()
        if f.is_file() and f.suffix.lower() in EXTENSOES_ACEITAS
    ]

    if not arquivos_brutos:
        logger.info("[%s] Nenhum arquivo em PENDENTES/", label)
        return resumo

    # ── 2. Detecta duplicatas (sigla repetida ou nome com '(N)') ──────────
    siglas_vistas: Dict[str, Path] = {}
    for arq in arquivos_brutos:
        if _eh_duplicata(arq.name):
            sys.exit(
                f"\n  ERRO: arquivo com nome de duplicata detectado em [{label}]:\n"
                f"     {arq.name}\n"
                f"  Renomeie ou apague antes de rodar de novo."
            )
        sigla = _normalizar_sigla(arq.name)
        if sigla in siglas_vistas:
            sys.exit(
                f"\n  ERRO: duas planilhas da mesma loja em [{label}]:\n"
                f"     {siglas_vistas[sigla].name}\n"
                f"     {arq.name}\n"
                f"  Apague/renomeie uma delas antes de rodar de novo."
            )
        siglas_vistas[sigla] = arq

    # ── 3. Abre a planilha pai (uma vez por região) ────────────────────────
    print(f"\n  [{label}] Processando {caminho_planilha_pai.name}")
    logger.info("[%s] Abrindo planilha pai: %s", label, caminho_planilha_pai)

    try:
        wb_pai = openpyxl.load_workbook(str(caminho_planilha_pai), keep_vba=True)
    except PermissionError:
        sys.exit(
            f"\n  ERRO: a planilha pai está aberta no Excel:\n"
            f"     {caminho_planilha_pai.name}\n"
            f"  Feche-a e tente novamente."
        )
    except Exception as exc:
        sys.exit(f"\n  ERRO ao abrir {caminho_planilha_pai.name}: {exc}")

    # ── 4. Processa cada arquivo ───────────────────────────────────────────
    mudou_alguma_coisa = False
    # Arquivos a mover para PROCESSADOS/ APÓS save bem-sucedido (não antes!)
    a_mover_processados: List[Tuple[Path, dict]] = []

    for sigla, arquivo in sorted(siglas_vistas.items()):
        # 4.1 Loja mapeada no YAML?
        if sigla not in cfg["lojas"]:
            logger.info("[%s] %s sem mapeamento no YAML — deixado em PENDENTES",
                        label, arquivo.name)
            print(f"     {sigla:<22} — sem mapeamento, deixado em PENDENTES")
            resumo["pendentes"].append({
                "loja": sigla,
                "arquivo": arquivo.name,
                "motivo": "Sem mapeamento no config_p2.yaml",
            })
            continue

        loja_cfg = cfg["lojas"][sigla]
        aba_pai = loja_cfg["aba_pai"]
        destinos = loja_cfg.get("destinos", {})

        # 4.2 Aba destino existe na planilha pai?
        if aba_pai not in wb_pai.sheetnames:
            mover_para_erros(
                arquivo, pasta_regiao, data_str, sigla,
                "ABA_PAI_AUSENTE",
                f"Aba '{aba_pai}' não existe na planilha pai",
                dry_run, logger,
            )
            resumo["com_erro"].append({
                "loja": sigla,
                "arquivo": arquivo.name,
                "motivo": "ABA_PAI_AUSENTE",
            })
            print(f"     {sigla:<22} — ERRO: aba '{aba_pai}' não existe na planilha pai")
            continue

        ws_pai = wb_pai[aba_pai]

        # 4.3 Detecta colunas USOU novas (criadas pelo Processo 1) nessa aba
        cols_usou_novas = encontrar_cols_usou_novas(ws_pai)
        if not cols_usou_novas:
            mover_para_erros(
                arquivo, pasta_regiao, data_str, sigla,
                "USOU_NAO_ENCONTRADO",
                f"Nenhuma coluna USOU encontrada em '{aba_pai}'. Processo 1 rodou?",
                dry_run, logger,
            )
            resumo["com_erro"].append({
                "loja": sigla, "arquivo": arquivo.name,
                "motivo": "USOU_NAO_ENCONTRADO",
            })
            print(f"     {sigla:<22} — ERRO: USOU não encontrado em '{aba_pai}'")
            continue

        # 4.4 Lê o arquivo da loja
        buffer, motivo, detalhe = ler_arquivo_loja(
            arquivo, cfg["aba_loja"], destinos, logger
        )

        if motivo == "ARQUIVO_ABERTO":
            # Pula, deixa em PENDENTES
            print(f"     {sigla:<22} — arquivo aberto no Excel, deixado em PENDENTES")
            resumo["pendentes"].append({
                "loja": sigla, "arquivo": arquivo.name, "motivo": motivo,
            })
            continue

        if motivo is not None:
            # Outros erros: move para ERROS/
            mover_para_erros(
                arquivo, pasta_regiao, data_str, sigla, motivo, detalhe,
                dry_run, logger,
            )
            resumo["com_erro"].append({
                "loja": sigla, "arquivo": arquivo.name, "motivo": motivo,
            })
            # Mostra o detalhe real (ex: mensagem do openpyxl) direto no console
            print(f"     {sigla:<22} — ERRO: {motivo}")
            if detalhe:
                print(f"       └─ {detalhe}")
            continue

        # 4.5 Verifica se as colunas USOU necessárias existem
        usous_necessarios = {d["usou"] for d in destinos.values()}
        usous_ausentes = usous_necessarios - set(cols_usou_novas.keys())
        if usous_ausentes:
            mover_para_erros(
                arquivo, pasta_regiao, data_str, sigla,
                "USOU_INSUFICIENTE",
                f"Aba '{aba_pai}' não tem USOU(s) {sorted(usous_ausentes)}. "
                f"Encontrados: {sorted(cols_usou_novas.keys())}",
                dry_run, logger,
            )
            resumo["com_erro"].append({
                "loja": sigla, "arquivo": arquivo.name, "motivo": "USOU_INSUFICIENTE",
            })
            print(f"     {sigla:<22} — ERRO: USOU(s) {sorted(usous_ausentes)} ausente(s)")
            continue

        # 4.6 Aplica o buffer (escreve nas células da planilha pai)
        celulas_escritas_usou1 = 0
        celulas_escritas_usou2 = 0
        celulas_escritas_usou3 = 0
        celulas_puladas = 0

        for celula_origem, destino in destinos.items():
            linha = destino["linha"]
            idx_usou = destino["usou"]
            col_destino = cols_usou_novas[idx_usou]

            if celula_origem not in buffer:
                # Célula origem estava vazia — pula
                celulas_puladas += 1
                continue

            valor = buffer[celula_origem]
            if not dry_run:
                ws_pai.cell(row=linha, column=col_destino).value = valor
            if idx_usou == 1:
                celulas_escritas_usou1 += 1
            elif idx_usou == 2:
                celulas_escritas_usou2 += 1
            elif idx_usou == 3:
                celulas_escritas_usou3 += 1

        mudou_alguma_coisa = True

        # 4.7 Acumula info para o resumo, mas NÃO move arquivo ainda
        # (move só APÓS save bem-sucedido da planilha pai — em caso de falha,
        # arquivo continua em PENDENTES para próxima execução)
        total_escritas = celulas_escritas_usou1 + celulas_escritas_usou2 + celulas_escritas_usou3
        partes = []
        if celulas_escritas_usou1:
            partes.append(f"USOU1={celulas_escritas_usou1}")
        if celulas_escritas_usou2:
            partes.append(f"USOU2={celulas_escritas_usou2}")
        if celulas_escritas_usou3:
            partes.append(f"USOU3={celulas_escritas_usou3}")
        info_usou = ", ".join(partes) if partes else "0 escritas"

        print(f"     {aba_pai:<22} — {total_escritas} valores escritos ({info_usou})")
        logger.info("[%s] %s: %d valores escritos (%s)", label, sigla, total_escritas, info_usou)

        info_processada = {
            "loja": sigla,
            "aba_pai": aba_pai,
            "arquivo": arquivo.name,
            "celulas_escritas": total_escritas,
            "celulas_escritas_usou1": celulas_escritas_usou1,
            "celulas_escritas_usou2": celulas_escritas_usou2,
            "celulas_escritas_usou3": celulas_escritas_usou3,
            "celulas_puladas_vazias": celulas_puladas,
        }
        a_mover_processados.append((arquivo, info_processada))

    # ── 5. Salva a planilha pai (se houve mudanças) ──────────────────────
    if mudou_alguma_coisa and not dry_run:
        try:
            wb_pai.save(str(caminho_planilha_pai))
            logger.info("[%s] Planilha pai salva: %s", label, caminho_planilha_pai.name)
        except PermissionError:
            wb_pai.close()
            sys.exit(
                f"\n  ERRO: não foi possível salvar {caminho_planilha_pai.name}.\n"
                f"  A planilha foi aberta no Excel durante a execução.\n"
                f"  Feche-a e rode novamente — os arquivos das lojas continuam em PENDENTES/."
            )
        except Exception as exc:
            wb_pai.close()
            sys.exit(f"\n  ERRO ao salvar {caminho_planilha_pai.name}: {exc}")
    elif dry_run and mudou_alguma_coisa:
        logger.info("[DRY-RUN] %s NÃO foi salvo (modo simulação)", caminho_planilha_pai.name)

    wb_pai.close()

    # ── 6. SÓ AGORA move os arquivos processados (save já confirmado) ────
    for arquivo, info_processada in a_mover_processados:
        destino_final = mover_para_processados(
            arquivo, pasta_regiao, data_str, dry_run, logger
        )
        info_processada["destino_arquivo"] = str(destino_final)
        resumo["processadas"].append(info_processada)

    return resumo


# ─────────────────────────────────────────────────────────────────────────────
# LOCK FILE
# ─────────────────────────────────────────────────────────────────────────────

def criar_lock(script_dir: Path) -> Path:
    lock_path = script_dir / LOCK_FILE_NAME
    if lock_path.exists():
        sys.exit(
            f"\n  ERRO: {LOCK_FILE_NAME} já existe — outra execução em andamento\n"
            f"  ou execução anterior travou.\n"
            f"  Se tiver certeza, apague o arquivo manualmente:\n"
            f"     {lock_path}\n"
        )
    lock_path.write_text(
        f"PID: criado em {dt.datetime.today().isoformat()}\n",
        encoding="utf-8",
    )
    return lock_path


def remover_lock(lock_path: Path) -> None:
    try:
        if lock_path.exists():
            lock_path.unlink()
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# VALIDAÇÃO: PROCESSO 1 RODOU ESTA SEMANA?
# ─────────────────────────────────────────────────────────────────────────────

_ABAS_IGNORAR_VALIDACAO = frozenset({
    "Planilha1", "AUXILIAR", "AUXILIAR VD'S",
    "LOJAS JAÚ", "LOJAS PRAIA", "LOJAS BAURU",
})


def validar_processo1_rodou(
    planilhas_pai: Dict[str, Path], logger: logging.Logger
) -> dt.date:
    """
    Abre cada planilha pai e detecta a data da última coluna USOU.
    Para performance, verifica APENAS a primeira aba elegível de cada planilha
    (Processo 1 é atômico por arquivo — ou rodou em todas as abas, ou em nenhuma).
    Aborta se a data NÃO for da semana atual.
    """
    hoje = dt.date.today()
    datas_encontradas: List[Tuple[str, dt.date]] = []

    for regiao, caminho in planilhas_pai.items():
        if not caminho.exists():
            logger.warning("Planilha pai %s não existe: %s", regiao, caminho)
            continue
        try:
            wb = openpyxl.load_workbook(str(caminho), data_only=True, read_only=False)
        except PermissionError:
            sys.exit(
                f"\n  ERRO: a planilha pai está aberta no Excel:\n"
                f"     {caminho.name}\n"
                f"  Feche-a e tente novamente."
            )
        except Exception as exc:
            sys.exit(f"\n  ERRO ao abrir {caminho.name}: {exc}")

        try:
            ultima_data: Optional[dt.date] = None
            for nome_aba in wb.sheetnames:
                if nome_aba in _ABAS_IGNORAR_VALIDACAO:
                    continue
                if nome_aba.startswith("__") or nome_aba.upper().startswith("TOTAL"):
                    continue
                ws = wb[nome_aba]
                d = detectar_data_processo1(ws)
                if d is not None:
                    ultima_data = d
                    break  # Suficiente: Processo 1 é atômico
            if ultima_data is None:
                logger.warning("Nenhuma data USOU encontrada em %s", caminho.name)
                continue
            datas_encontradas.append((regiao, ultima_data))
        finally:
            wb.close()

    if not datas_encontradas:
        sys.exit(
            "\n  ERRO: nenhuma data encontrada nas colunas USOU das planilhas pai.\n"
            "  O Processo 1 precisa rodar antes do Processo 2."
        )

    erros = [(r, d) for r, d in datas_encontradas if not _mesma_semana(d, hoje)]
    if erros:
        msg = "\n".join(
            f"     {r.upper()}: última coluna USOU é de {d.strftime('%d/%m/%Y')}"
            for r, d in erros
        )
        sys.exit(
            f"\n  ERRO: Processo 1 não rodou esta semana em todas as regiões:\n"
            f"{msg}\n"
            f"  Hoje é {hoje.strftime('%d/%m/%Y')}.\n"
            f"  Rode o Processo 1 antes de continuar."
        )

    return max(d for _, d in datas_encontradas)


# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────

def _setup_logging(log_dir: Path, dry_run: bool) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    hoje = dt.date.today().strftime("%Y-%m-%d")
    sufixo = "_dry-run" if dry_run else ""
    log_path = log_dir / f"processo2_{hoje}{sufixo}.log"

    logger = logging.getLogger("processo2")
    logger.setLevel(logging.DEBUG)

    fh = logging.FileHandler(str(log_path), encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stderr)
    ch.setLevel(logging.WARNING)
    ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(ch)

    return logger


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    # UTF-8 no Windows
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(
        description="Processo 2 — Integração das planilhas das lojas nas Planilhas Pai"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Simula sem mover arquivos nem salvar planilhas")
    parser.add_argument("--regiao", choices=["praia", "jau", "bauru"],
                        help="Processar apenas uma região")
    args = parser.parse_args()

    script_dir = Path(__file__).parent

    # Carrega config
    config_path = script_dir / "config_p2.yaml"
    if not config_path.exists():
        sys.exit(f"ERRO: config_p2.yaml não encontrado em {script_dir}")
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    def resolve(p: str) -> Path:
        return (script_dir / p).resolve()

    logger = _setup_logging(resolve(cfg["dirs"]["logs"]), args.dry_run)

    # Cabeçalho
    hoje_fmt = dt.date.today().strftime("%d/%m/%Y")
    dry_tag = " [DRY-RUN]" if args.dry_run else ""
    print("=" * 60)
    print(f"  PROCESSO 2 — ANTILHAS — {hoje_fmt}{dry_tag}")
    print("=" * 60)
    # Diagnóstico: mostra qual Python e openpyxl estão de fato em uso
    print(f"   Python:    {sys.executable}")
    print(f"   openpyxl:  {openpyxl.__version__}  ({Path(openpyxl.__file__).parent})")
    logger.info("Iniciando Processo 2%s", dry_tag)
    logger.info("Python: %s  |  openpyxl: %s  (%s)",
                sys.executable, openpyxl.__version__,
                Path(openpyxl.__file__).parent)

    # Trava
    lock_path = criar_lock(script_dir)
    inicio = time.time()

    try:
        # Resolve caminhos das planilhas pai
        planilhas_pai_paths = {
            r: resolve(p) for r, p in cfg["planilhas_pai"].items()
        }

        # Filtra por região se solicitado
        if args.regiao:
            planilhas_pai_paths = {args.regiao: planilhas_pai_paths[args.regiao]}

        # Valida que Processo 1 rodou esta semana
        print("   Verificando se o Processo 1 rodou esta semana...")
        data_p1 = validar_processo1_rodou(planilhas_pai_paths, logger)
        print(f"   OK — última coluna USOU em {data_p1.strftime('%d/%m/%Y')}")

        # Backup
        fazer_backup(
            planilhas_pai_paths,
            resolve(cfg["dirs"]["backup"]),
            args.dry_run, logger,
        )

        # Processa cada região
        dir_entrada = resolve(cfg["dir_entrada"])
        data_str = dt.date.today().strftime("%Y-%m-%d")
        mapeamento_pasta = {"praia": "PRAIA", "jau": "JAU", "bauru": "BAURU"}
        resumo_regioes: dict = {}

        for regiao, caminho_pai in planilhas_pai_paths.items():
            pasta_regiao = dir_entrada / mapeamento_pasta[regiao]
            resumo = processar_regiao(
                regiao, pasta_regiao, caminho_pai, cfg,
                args.dry_run, data_str, logger,
            )
            resumo_regioes[regiao] = resumo

        # Manifesto
        total_processadas = sum(len(r["processadas"]) for r in resumo_regioes.values())
        total_erros = sum(len(r["com_erro"]) for r in resumo_regioes.values())
        total_pendentes = sum(len(r["pendentes"]) for r in resumo_regioes.values())

        manifesto = {
            "data_execucao": data_str,
            "horario_inicio": dt.datetime.fromtimestamp(inicio).strftime("%H:%M:%S"),
            "horario_fim": dt.datetime.today().strftime("%H:%M:%S"),
            "duracao_segundos": round(time.time() - inicio, 2),
            "dry_run": args.dry_run,
            "regiao_filtro": args.regiao,
            "processo1_data_ultima_coluna": data_p1.strftime("%Y-%m-%d"),
            "totais": {
                "processadas": total_processadas,
                "com_erro": total_erros,
                "pendentes": total_pendentes,
            },
            "regioes": resumo_regioes,
        }

        if not args.dry_run:
            log_dir = resolve(cfg["dirs"]["logs"])
            log_dir.mkdir(parents=True, exist_ok=True)
            manifesto_path = log_dir / f"processo2_{data_str}.json"
            with open(manifesto_path, "w", encoding="utf-8") as f:
                json.dump(manifesto, f, indent=2, ensure_ascii=False)
            logger.info("Manifesto gerado: %s", manifesto_path)

        # Resumo final no console
        print()
        print("  Resumo:")
        print(f"     Processadas: {total_processadas} loja(s)")
        print(f"     Com erro:    {total_erros}")
        print(f"     Pendentes:   {total_pendentes}")

        if not args.dry_run and total_processadas > 0:
            print(f"\n   Arquivos movidos para: PROCESSADOS/{data_str}/")

        print("\n   Processo 2 concluído com sucesso.")
        print("=" * 60)
        logger.info("Processo 2 concluído.")

    finally:
        remover_lock(lock_path)


if __name__ == "__main__":
    main()
